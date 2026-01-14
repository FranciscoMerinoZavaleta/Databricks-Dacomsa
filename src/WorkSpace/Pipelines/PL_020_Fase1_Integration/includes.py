# Databricks notebook source
# MAGIC %md
# MAGIC #Includes

# COMMAND ----------

# MAGIC %md
# MAGIC ##Librerias

# COMMAND ----------

import os
import zipfile
import xml.etree.ElementTree as ET
import json
import requests
import json
import time
from datetime import datetime
import unicodedata


# COMMAND ----------

#se coloca porque falta instalarse marca cuando se ocupa correr hoja de rebate
!pip install openpyxl


# COMMAND ----------

import openpyxl

# COMMAND ----------

from pyspark.sql.functions import *
from pyspark.sql import Window
from pyspark.sql.types import StructType,StructField, StringType, IntegerType, DoubleType, DateType, TimestampType, BooleanType, DecimalType,TimestampType
import numpy as np
import os, uuid, sys

# COMMAND ----------

from pyspark.sql import SparkSession, Row
# from pyspark.sql.functions import (date_format, col, regexp_extract, length, when, lit, lead, collect_list, monotonically_increasing_id, lag, last, sum, coalesce, concat_ws, concat, format_number, current_date, date_add, expr,add_months)
from pyspark.sql.types import IntegerType, StringType, StringType, DateType
from pyspark.sql.window import Window
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

# COMMAND ----------

!pip install unidecode
from unidecode import unidecode

# COMMAND ----------

!pip install unidecode
from unidecode import unidecode

# COMMAND ----------

!pip install pmdarima
import pmdarima as pm
from pmdarima.arima import auto_arima

# COMMAND ----------

!pip install --upgrade numpy scipy statsmodels prophet
import statsmodels.api as sm
from prophet import Prophet
from sklearn.metrics import mean_absolute_percentage_error, mean_absolute_error
import logging

# COMMAND ----------

import warnings
warnings.filterwarnings('ignore')
import pyspark.sql.functions as F

# COMMAND ----------

import pandas as pd
import numpy as np
import pyspark.pandas as ps
#import geopandas as gpd
#from shapely.geometry import Point

# COMMAND ----------

# MAGIC %md
# MAGIC ##Funciones

# COMMAND ----------

def lsR(path,last_modified ): #**kwargs):
    #last_modified=kwargs.get('last_modified', None)

    # if (last_modified == None):
    #     last_modified=pd.Timestamp('2023-11-08 00:00:00', tz=None)
    
    return {
        fname.replace('dbfs:', '').replace('/_SUCCESS', '') 
        for flist in [
            ([fi.path+'*'+str(fi.modificationTime)] if fi.isFile() else lsR(fi.path, last_modified))
            #([pd.to_datetime(int(str(row.modificationTime)+'000000') )] if fi.isFile()  else lsR(fi.path))
            for fi in dbutils.fs.ls(path) 
            if not(fi.isFile()) or pd.to_datetime(int(str(fi.modificationTime)+'000000')) > last_modified  
        ]
        for fname in flist
    }

# COMMAND ----------

##Esquema tablas SQL
def get_schema(table=False):
    schema_dict = {"FACT_IND_SD003_MARGEN_CLIENTE_PRODUCTO" : StructType([StructField('ANIO_MES', StringType(), True),
                                StructField('ID_COMPANIA', StringType(), True),
                                StructField('COMPANIA', StringType(), True),
                                StructField('ID_SISTEMA', StringType(), True),
                                StructField('SISTEMA', StringType(), True),
                                StructField('ID_LINEA_ABIERTA', StringType(), True),
                                StructField('LINEA_ABIERTA', StringType(), True),
                                StructField('LINEA_AGRUPADA', StringType(), True),
                                StructField('MERCADO', StringType(), True),
                                StructField('GRUPO_CLIENTE', StringType(), True),
                                StructField('ID_CLIENTE', StringType(), True),
                                StructField('CLIENTE', StringType(), True),
                                StructField('UNIDAD_MEDIDA', StringType(), True),
                                StructField('MATNR', StringType(), True),
                                StructField('MONEDA', StringType(), True),
                                StructField('TIPO_CAMBIO', DecimalType(), True),
                                StructField('CANTIDAD', DoubleType(), True),
                                StructField('VENTA TOTAL', DoubleType(), True),
                                StructField('COSTO_VENTA_FACT', DoubleType(), True),
                                StructField('MARGEN_CLIENTE', DoubleType(), True),
                                StructField('MARGEN_CLIENTE_PRC', DoubleType(), True),
                                StructField('COSTO_UNIT_FABRICACION', DoubleType(), True),
                                StructField('COSTO_VENTA_STD_FABR', DoubleType(), True),
                                StructField('COSTO_VENTA_PRECIO_STD_PLANTA', DoubleType(), True),
                                StructField('COSTO_VENTA_PRECIO_VAR_PLANTA', DoubleType(), True),
                                StructField('COSTO_VENTA_PRECIO_STD_VAR_PLANTA', DoubleType(), True),
                                StructField('VENTA_TOTAL_MONEDA_FACT', DoubleType(), True),
                                StructField('TOTAL_VENTA_CABECERO_FACT', DoubleType(), True),
                                StructField('MATERIA_PRIMA', DoubleType(), True),
                                StructField('SEMITERMINADO', DoubleType(), True),
                                StructField('MANO_OBRA', DoubleType(), True),
                                StructField('MAQUILA', DoubleType(), True),
                                StructField('DEPRECIACION', DoubleType(), True),
                                StructField('TIEMPO_PREPARACION', DoubleType(), True),
                                StructField('GIF', DoubleType(), True),
                                StructField('HER_SEMITERMINADO', DoubleType(), True),
                                StructField('HER_PROTERMINADO', DoubleType(), True),
                                StructField('GIF_ABS', DoubleType(), True),
                                StructField('GDF_LUZ', DoubleType(), True),
                                StructField('GDF_COMBUSTIBLE', DoubleType(), True),
                                StructField('GDF_OTROS', DoubleType(), True),
                                StructField('GIF_MANOBRA', DoubleType(), True),
                                StructField('GIF_MANTENIMIENTO', DoubleType(), True),
                                StructField('GIF_OTROS', DoubleType(), True),
                                StructField('GIF_SERVICIO', DoubleType(), True),
                                StructField('GDF', DoubleType(), True),
                                StructField('PRODUCTO_TERMINADO', DoubleType(), True),
                                StructField('MERMA_MAQUILA', DoubleType(), True),
                                StructField('FLAG_CARGA_SAC', IntegerType(), True),
                                StructField('FECCARGA', TimestampType(), True),
                                StructField('USRCARGA', StringType(), True),
                                StructField('FECHA_INFORMACION', StringType(), True)]),
        "FACT_IND_OP006_HAB_INV" : StructType([StructField('ANIO_MES', StringType(), True),
                                StructField('ID_COMPANIA', StringType(), True),
                                StructField('COMPANIA', StringType(), True),
                                StructField('ID_SISTEMA', StringType(), True),
                                StructField('SISTEMA', StringType(), True),
                                StructField('ID_LINEA_ABIERTA', StringType(), True),
                                StructField('LINEA_ABIERTA', StringType(), True),
                                StructField('TIPO_MATERIAL', StringType(), True),
                                StructField('IND_ABC', StringType(), True),
                                StructField('ID_MATERIAL', StringType(), True),
                                StructField('MATERIAL', StringType(), True),
                                StructField('HABILIDAD_ITEM', StringType(), True),
                                StructField('STOCK_UNIDADES', DoubleType(), True),
                                StructField('STOCK_VALOR', DoubleType(), True),
                                StructField('PRONOSTICO_UNIDADES', DoubleType(), True),
                                StructField('PRONOSTICO_VALOR', DoubleType(), True),
                                StructField('VENTA_UNIDADES', DoubleType(), True),
                                StructField('VENTA_IMPORTE', DoubleType(), True),
                                StructField('VENTA_IMPORTE_REAL', DoubleType(), True),
                                StructField('PEDIDO_UNIDADES', DoubleType(), True),
                                StructField('PEDIDO_IMPORTE', DoubleType(), True),
                                StructField('PEDIDO_IMPORTE_REAL', DoubleType(), True),
                                StructField('PEDIDO_CANCELADO_UNIDADES', DoubleType(), True),
                                StructField('PEDIDO_CANCELADO_IMPORTE', DoubleType(), True),
                                StructField('PEDIDO_CANCELADO_IMPORTE_REAL', DoubleType(), True),
                                StructField('PEDIDO_RETENIDO_UNIDADES', DoubleType(), True),
                                StructField('PEDIDO_RETENIDO_IMPORTE', DoubleType(), True),
                                StructField('PEDIDO_RETENIDO_IMPORTE_REAL', DoubleType(), True),
                                StructField('IND_MATERIAL_VENTA', IntegerType(), True),
                                StructField('ORIGEN_PRONOSTICO', StringType(), True),
                                StructField('FECINFORMACION', DateType(), True),
                                StructField('FECCARGA', TimestampType(), True),
                                StructField('FECULTACT', TimestampType(), True),
                                StructField('USRCARGA', StringType(), True),
                                StructField('ID_CENTRO', StringType(), True),
                                StructField('CENTRO', StringType(), True),
                                StructField('FLAG_CS', StringType(), True),
                                StructField('FLAG_CARGA_SAC', IntegerType(), True)]),
        "FACT_IND_SD_FILLRATE" : StructType([StructField('ANIO_MES', StringType(), True),
                                StructField('PERIODO', DateType(), True),
                                StructField('ID_PLANTA', StringType(), True),
                                StructField('PLANTA', StringType(), True),
                                StructField('ID_SISTEMA', StringType(), True),
                                StructField('SISTEMA', StringType(), True),
                                StructField('LINEA_AGRUPADA', StringType(), True),
                                StructField('ID_CLIENTE', StringType(), True),
                                StructField('CLIENTE', StringType(), True),
                                StructField('ID_INDICADOR_ABC', StringType(), True),
                                StructField('NUM_ITEMS', IntegerType(), True),
                                StructField('CANT_PEDIDO', DoubleType(), True),
                                StructField('CANT_FACTURA', DoubleType(), True),
                                StructField('CANT_RETENIDA', DoubleType(), True),
                                StructField('CANT_FACTURA_AJUSTADA', DoubleType(), True),
                                StructField('IMP_PEDIDO', DoubleType(), True),
                                StructField('IMP_FACTURA', DoubleType(), True),
                                StructField('IMP_RETENIDO', DoubleType(), True),
                                StructField('IMP_FACTURA_AJUSTADA', DoubleType(), True),
                                StructField('POS_SOLICITADAS', DoubleType(), True),
                                StructField('POS_ENTREGADAS', DoubleType(), True),
                                StructField('FECCARGA', TimestampType(), True),
                                StructField('FECULTACT', TimestampType(), True),
                                StructField('USRCARGA', StringType(), True),
                                StructField('FECINFORMACION', DateType(), True),
                                StructField('FLAG_CARGA_SAC', IntegerType(), True)]),
         "FACT_PEDIDOS_ESTATUS" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_PEDIDO_VENTA', StringType(), True),
                                StructField('POS_PEDIDO_VENTA', DoubleType(), True),
                                StructField('FECHA_REGISTRO', TimestampType(), True),
                                StructField('FECHA_CANCELACION', TimestampType(), True),
                                StructField('FECHA_PEDIDO', TimestampType(), True),
                                StructField('PIEZAS', DoubleType(), True),
                                StructField('ID_MONEDA', StringType(), True),
                                StructField('ID_CLIENTE_SOLICITANTE', StringType(), True),
                                StructField('ID_PERIODO', StringType(), True),
                                StructField('ID_MATERIAL', StringType(), True),
                                StructField('ID_MOTIVO_CANCELACION_PEDIDO', StringType(), True),
                                StructField('ID_SOCIEDAD_FACTURADORA', StringType(), True),
                                StructField('ID_CLASE_DOCUMENTO_VENTA', StringType(), True),
                                StructField('ID_CANAL_DISTRIBUCION', StringType(), True),
                                StructField('ESTATUS_PEDIDO', StringType(), True),
                                StructField('FECHA_COMPROMISO', TimestampType(), True),
                                StructField('PIEZAS_FACTURADA', DoubleType(), True),
                                StructField('PIEZAS_PENDIENTE', DoubleType(), True),
                                StructField('PIEZAS_PICKING', DoubleType(), True),
                                StructField('PIEZAS_BLOQUEADA', DoubleType(), True),
                                StructField('PIEZAS_RECHAZADA', DoubleType(), True),
                                StructField('PIEZAS_BACKORDER', DoubleType(), True),
                                StructField('PRECIO_NETO', DoubleType(), True),
                                StructField('IMPORTE_FACTURADA', DoubleType(), True),
                                StructField('IMPORTE_PICKING', DoubleType(), True),
                                StructField('IMPORTE_PENDIENTE', DoubleType(), True),
                                StructField('IMPORTE_BLOQUEADA', DoubleType(), True),
                                StructField('IMPORTE_RECHAZADA', DoubleType(), True),
                                StructField('IMPORTE_BACKORDER', DoubleType(), True),
                                StructField('IMPORTE', DoubleType(), True),
                                StructField('FECHA_FACTURA', TimestampType(), True),
                                StructField('TIPO_CAMBIO_DPRECIOS', DoubleType(), True),
                                StructField('ID_ESTATUS_VERIFICACION_CREDITO', StringType(), True)]),
          "FACT_PEDIDOS_ESTATUS_ULTIMO_REGISTRO" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_PEDIDO_VENTA', StringType(), True),
                                StructField('POS_PEDIDO_VENTA', DoubleType(), True),
                                StructField('FECHA_REGISTRO', TimestampType(), True),
                                StructField('FECHA_CANCELACION', TimestampType(), True),
                                StructField('FECHA_PEDIDO', TimestampType(), True),
                                StructField('PIEZAS', DoubleType(), True),
                                StructField('ID_MONEDA', StringType(), True),
                                StructField('ID_CLIENTE_SOLICITANTE', StringType(), True),
                                StructField('ID_PERIODO', StringType(), True),
                                StructField('ID_MATERIAL', StringType(), True),
                                StructField('ID_MOTIVO_CANCELACION_PEDIDO', StringType(), True),
                                StructField('ID_SOCIEDAD_FACTURADORA', StringType(), True),
                                StructField('ID_CLASE_DOCUMENTO_VENTA', StringType(), True),
                                StructField('ID_CANAL_DISTRIBUCION', StringType(), True),
                                StructField('ESTATUS_PEDIDO', StringType(), True),
                                StructField('FECHA_COMPROMISO', TimestampType(), True),
                                StructField('PIEZAS_FACTURADA', DoubleType(), True),
                                StructField('PIEZAS_PENDIENTE', DoubleType(), True),
                                StructField('PIEZAS_PICKING', DoubleType(), True),
                                StructField('PIEZAS_BLOQUEADA', DoubleType(), True),
                                StructField('PIEZAS_RECHAZADA', DoubleType(), True),
                                StructField('PIEZAS_BACKORDER', DoubleType(), True),
                                StructField('PRECIO_NETO', DoubleType(), True),
                                StructField('IMPORTE_FACTURADA', DoubleType(), True),
                                StructField('IMPORTE_PICKING', DoubleType(), True),
                                StructField('IMPORTE_PENDIENTE', DoubleType(), True),
                                StructField('IMPORTE_BLOQUEADA', DoubleType(), True),
                                StructField('IMPORTE_RECHAZADA', DoubleType(), True),
                                StructField('IMPORTE_BACKORDER', DoubleType(), True),
                                StructField('IMPORTE', DoubleType(), True),
                                StructField('FECHA_FACTURA', TimestampType(), True),
                                StructField('TIPO_CAMBIO_DPRECIOS', DoubleType(), True),
                                StructField('ID_ESTATUS_VERIFICACION_CREDITO', StringType(), True)]),
          "DIM_CLIENTE_VENTAS" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_ORGANIZACION_VENTAS', StringType(), True),
                                StructField('ID_CANAL_DISTRIBUCION', StringType(), True),
                                StructField('ID_SECTOR', StringType(), True),
                                StructField('ID_CLIENTE', StringType(), True),
                                StructField('NOMBRE_1', StringType(), True),
                                StructField('NOMBRE_2', StringType(), True),
                                StructField('DES_CLIENTE_CORTA', StringType(), True),
                                StructField('ID_GRUPO_CUENTAS', StringType(), True),
                                StructField('ID_DISTRITO_NIELSEN', StringType(), True),
                                StructField('ID_RAMO_1', StringType(), True),
                                StructField('CIUDAD', StringType(), True),
                                StructField('COLONIA', StringType(), True),
                                StructField('ID_OFICINA_VENTAS', StringType(), True),
                                StructField('ID_ZONA_VENTAS', StringType(), True),
                                StructField('RFC', StringType(), True),
                                StructField('ID_PAIS', StringType(), True),
                                StructField('ID_ESTADO', StringType(), True),
                                StructField('CALLE', StringType(), True),
                                StructField('NUMERO_CASA', StringType(), True),
                                StructField('CODIGO_POSTAL', StringType(), True),
                                StructField('ORDENADOR1', StringType(), True),
                                StructField('ORDENADOR2', StringType(), True),
                                StructField('ID_GRUPO_CLIENTES', StringType(), True),
                                StructField('ID_GRUPO_PRECIOS', StringType(), True),
                                StructField('ID_LISTA_PRECIOS', StringType(), True),
                                StructField('ID_CENTRO_ENTREGA', StringType(), True),
                                StructField('ID_GRUPO_VENTAS', StringType(), True),
                                StructField('ID_GRUPO_CLIENTES_1', StringType(), True),
                                StructField('ID_IDIOMA', StringType(), True),
                                StructField('ID_CLIENTE_AGRUPACION', StringType(), True),
                                StructField('DES_CLIENTE_AGRUPACION', StringType(), True),
                                StructField('ID_CLIENTE_SUPERVISOR', IntegerType(), True),
                                StructField('DES_CLIENTE_SUPERVISOR', StringType(), True),
                                StructField('ID_CLIENTE_PROMOTOR', IntegerType(), True),
                                StructField('DES_CLIENTE_PROMOTOR', StringType(), True),
                                StructField('ID_CLIENTE_SUBGRUPO', StringType(), True),
                                StructField('DES_CLIENTE_SUBGRUPO', StringType(), True),
                                StructField('NUM_CLIENTE', StringType(), True)]),
           "DIM_GRUPO_VENTAS" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_OFICINA_VENTAS', StringType(), True),
                                StructField('ID_GRUPO_VENTAS', StringType(), True),
                                StructField('DES_GRUPO_VENTAS', StringType(), True)]),
           "DIM_ESTADOS" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_PAIS', StringType(), True),
                                StructField('ID_ESTADO', StringType(), True),
                                StructField('DES_ESTADO', StringType(), True)]),
            "DIM_MATERIAL_VENTAS" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_MATERIAL', StringType(), True),
                                StructField('ID_CENTRO', StringType(), True),
                                StructField('ID_ORGANIZACION_VENTAS', StringType(), True),
                                StructField('ID_CANAL_DISTRIBUCION', StringType(), True),
                                StructField('ID_SECTOR', StringType(), True),
                                StructField('ID_TIPO_MATERIAL', StringType(), True),
                                StructField('ID_GRUPO_ARTICULO', StringType(), True),
                                StructField('ID_UNIDAD_MEDIDA_BASE', StringType(), True),
                                StructField('PESO_BRUTO', DoubleType(), True),
                                StructField('PESO_NETO', DoubleType(), True),
                                StructField('ID_UNIDAD_PESO', StringType(), True),
                                StructField('VOLUMEN', DoubleType(), True),
                                StructField('ID_UNIDAD_VOLUMEN', StringType(), True),
                                StructField('ID_JERARQUIA_PRODUCTO_AREA_VENTA', StringType(), True),
                                StructField('ESTATUS_ACTUALIZACION', StringType(), True),
                                StructField('ESTATUS_DISTRIBUCION', StringType(), True),
                                StructField('DES_MATERIAL', StringType(), True),
                                StructField('ID_GRUPO_MATERIAL', StringType(), True),
                                StructField('DES_GRUPO_MATERIAL', StringType(), True),
                                StructField('ID_GRUPO_MATERIAL_1', StringType(), True),
                                StructField('DES_GRUPO_MATERIAL_1', StringType(), True),
                                StructField('ID_GRUPO_MATERIAL_2', StringType(), True),
                                StructField('DES_GRUPO_MATERIAL_2', StringType(), True),
                                StructField('ID_GRUPO_MATERIAL_3', StringType(), True),
                                StructField('DES_GRUPO_MATERIAL_3', StringType(), True),
                                StructField('ID_GRUPO_MATERIAL_4', StringType(), True),
                                StructField('DES_GRUPO_MATERIAL_4', StringType(), True),
                                StructField('ID_GRUPO_MATERIAL_5', StringType(), True),
                                StructField('DES_GRUPO_MATERIAL_5', StringType(), True),
                                StructField('NUMERADOR_UNIDAD_CONV_BASE', DoubleType(), True),
                                StructField('DENOMINADOR_UNIDAD_CONV_BASE', DoubleType(), True),
                                StructField('ID_ESTATUS_VENTAS', StringType(), True),
                                StructField('DES_JERARQUIA_PRODUCTO_AREA_VENTA', StringType(), True),
                                StructField('PRECIO_MEDIO_VARIABLE', DoubleType(), True),
                                StructField('PRECIO_ESTANDAR', DoubleType(), True),
                                StructField('ID_INDICADOR_ABC', StringType(), True),
                                StructField('DES_GRUPO_ARTICULO', StringType(), True),
                                StructField('DES_LARGA_GRUPO_ARTICULO', StringType(), True),
                                StructField('INDICADOR_CRITICO', StringType(), True),
                                StructField('DES_PLANIFICABLE', StringType(), True),
                                StructField('DES_ESTATUS_VENTAS', StringType(), True),
                                StructField('DES_SECTOR', StringType(), True),
                                StructField('NUM_MATERIAL', StringType(), True)]),
            "DIM_ALMACEN" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_SOCIEDAD', StringType(), True),
                                StructField('ID_CENTRO', StringType(), True),
                                StructField('ID_ALMACEN', StringType(), True),
                                StructField('DES_ALMACEN', StringType(), True),
                                StructField('ID_SECTOR', StringType(), True)]),
            "DIM_CANAL_DISTRIBUCION" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_CANAL_DISTRIBUCION', StringType(), True),
                                StructField('DES_CANAL_DISTRIBUCION', StringType(), True)]),
            "DIM_CENTRO" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_SOCIEDAD', StringType(), True),
                                StructField('ID_CENTRO', StringType(), True),
                                StructField('DES_CENTRO', StringType(), True)]),
            "DIM_ESCENARIO" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_ESCENARIO', StringType(), True),
                                StructField('DESC_ESCENARIO', StringType(), True)]),     
            "DIM_PPT_CLIENTE" : StructType([StructField('SISTEMA', StringType(), True),
                                StructField('ID_GRUPO', StringType(), True),
                                StructField('GRUPO', StringType(), True),
                                StructField('SHIP_TO', StringType(), True),
                                StructField('ID_CLIENTE', StringType(), True),
                                StructField('DESC_CLIENTE', StringType(), True),
                                StructField('ESTADO', StringType(), True),
                                StructField('ID_ASESOR', StringType(), True),
                                StructField('DESC_REGIONAL', StringType(), True),
                                StructField('DESC_ASESOR', StringType(), True),
                                StructField('LINEA MOD', StringType(), True),
                                StructField('MERCADO', StringType(), True),
                                StructField('FECHA', TimestampType(), True),
                                StructField('IMPORTE', DoubleType(), True),
                                StructField('INDICE', DoubleType(), True),
                                StructField('FECHA_CARGA', TimestampType(), True),
                                StructField('TIPO_CARGA',BooleanType(), True),
                                StructField('PIEZAS', DecimalType(), True)]),
             "DIM_ZONA_VENTAS" : StructType([StructField('ID_MANDANTE', StringType(), True),
                                StructField('ID_ZONA_VENTAS', StringType(), True),
                                StructField('DES_ZONA_VENTAS', StringType(), True)]),
                      "CTL_REBATE" : StructType([StructField('AÑO', IntegerType(), True),
                                StructField('ID_GRUPO', DoubleType(), True),
                                StructField('GRUPO', StringType(), True),
                                StructField('ID_CLIENTE', StringType(), True),
                                StructField('CLIENTE', StringType(), True),
                                StructField('LINEA', StringType(), True),
                                StructField('OBJETIVO', DoubleType(), True),
                                StructField('REBATE_NORMAL', DoubleType(), True),
                                StructField('TIPO_REBATE_NORMAL', StringType(), True),
                                StructField('REBATE_ESPECIAL', StringType(), True),
                                StructField('TIPO_REBATE_ESPECIAL', StringType(), True),
                                StructField('RESPONSABLE', StringType(), True),
                                StructField('NOTAS', StringType(), True)]),
             "CTL_REBATE_CLIENTE" : StructType([StructField('AÑO', DoubleType(), True),
                                StructField('ID_CLIENTE', StringType(), True),
                                StructField('CLIENTE', StringType(), True),
                                StructField('LINEA', StringType(), True),
                                StructField('OBJETIVO', DoubleType(), True),
                                StructField('REBATE_NORMAL', DoubleType(), True),
                                StructField('TIPO_REBATE_NORMAL', StringType(), True),
                                StructField('REBATE_ESPECIAL', StringType(), True),
                                StructField('TIPO_REBATE_ESPECIAL', StringType(), True),
                                StructField('RESPONSABLE', StringType(), True),
                                StructField('NOTAS', StringType(), True)]),
                "TMP_IND_ECC_MARA" : StructType([StructField('MANDT', StringType(), True),
                                StructField('MATNR', StringType(), True),
                                StructField('ERSDA', DateType(), True),
                                StructField('ERNAM', StringType(), True),
                                StructField('LAEDA', DateType(), True),
                                StructField('AENAM', StringType(), True),
                                StructField('VPSTA', StringType(), True),
                                StructField('PSTAT', StringType(), True),
                                StructField('LVORM', StringType(), True),
                                StructField('MTART', StringType(), True),
                                StructField('MBRSH', StringType(), True),
                                StructField('MATKL', StringType(), True),
                                StructField('BISMT', StringType(), True),
                                StructField('MEINS', StringType(), True),
                                StructField('BSTME', StringType(), True),
                                StructField('ZEINR', StringType(), True),
                                StructField('ZEIAR', StringType(), True),
                                StructField('ZEIVR', StringType(), True),
                                StructField('ZEIFO', StringType(), True),
                                StructField('AESZN', StringType(), True),
                                StructField('BLATT', StringType(), True),
                                StructField('BLANZ', DoubleType(), True),
                                StructField('FERTH', StringType(), True),
                                StructField('FORMT', StringType(), True),
                                StructField('GROES', StringType(), True),
                                StructField('WRKST', StringType(), True),
                                StructField('NORMT', StringType(), True),
                                StructField('LABOR', StringType(), True),
                                StructField('EKWSL', StringType(), True),
                                StructField('BRGEW', DoubleType(), True),
                                StructField('NTGEW', DoubleType(), True),
                                StructField('GEWEI', StringType(), True),
                                StructField('VOLUM', DoubleType(), True),
                                StructField('VOLEH', StringType(), True),
                                StructField('BEHVO', StringType(), True),
                                StructField('RAUBE', StringType(), True),
                                StructField('TEMPB', StringType(), True),
                                StructField('DISST', StringType(), True),
                                StructField('TRAGR', StringType(), True),
                                StructField('STOFF', StringType(), True),
                                StructField('SPART', StringType(), True),
                                StructField('KUNNR', StringType(), True),
                                StructField('EANNR', StringType(), True),
                                StructField('WESCH', DoubleType(), True),
                                StructField('BWVOR', StringType(), True),
                                StructField('BWSCL', StringType(), True),
                                StructField('SAISO', StringType(), True),
                                StructField('ETIAR', StringType(), True),
                                StructField('ETIFO', StringType(), True),
                                StructField('ENTAR', StringType(), True),
                                StructField('EAN11', StringType(), True),
                                StructField('NUMTP', StringType(), True),
                                StructField('LAENG', DoubleType(), True),
                                StructField('BREIT', DoubleType(), True),
                                StructField('HOEHE', DoubleType(), True),
                                StructField('MEABM', StringType(), True),
                                StructField('PRDHA', StringType(), True),
                                StructField('AEKLK', StringType(), True),
                                StructField('CADKZ', StringType(), True),
                                StructField('QMPUR', StringType(), True),
                                StructField('ERGEW', DoubleType(), True),
                                StructField('ERGEI', StringType(), True),
                                StructField('ERVOL', DoubleType(), True),
                                StructField('ERVOE', StringType(), True),
                                StructField('GEWTO', DoubleType(), True),
                                StructField('VOLTO', DoubleType(), True),
                                StructField('VABME', StringType(), True),
                                StructField('KZREV', StringType(), True),
                                StructField('KZKFG', StringType(), True),
                                StructField('XCHPF', StringType(), True),
                                StructField('VHART', StringType(), True),
                                StructField('FUELG', DoubleType(), True),
                                StructField('STFAK', IntegerType(), True),
                                StructField('MAGRV', StringType(), True),
                                StructField('BEGRU', StringType(), True),
                                StructField('DATAB', DateType(), True),
                                StructField('LIQDT', DateType(), True),
                                StructField('SAISJ', StringType(), True),
                                StructField('PLGTP', StringType(), True),
                                StructField('MLGUT', StringType(), True),
                                StructField('EXTWG', StringType(), True),
                                StructField('SATNR', StringType(), True),
                                StructField('ATTYP', StringType(), True),
                                StructField('KZKUP', StringType(), True),
                                StructField('KZNFM', StringType(), True),
                                StructField('PMATA', StringType(), True),
                                StructField('MSTAE', StringType(), True),
                                StructField('MSTAV', StringType(), True),
                                StructField('MSTDE', DateType(), True),
                                StructField('MSTDV', DateType(), True),
                                StructField('TAKLV', StringType(), True),
                                StructField('RBNRM', StringType(), True),
                                StructField('MHDRZ', DoubleType(), True),
                                StructField('MHDHB', DoubleType(), True),
                                StructField('MHDLP', DoubleType(), True),
                                StructField('INHME', StringType(), True),
                                StructField('INHAL', DoubleType(), True),
                                StructField('VPREH', DoubleType(), True),
                                StructField('ETIAG', StringType(), True),
                                StructField('INHBR', DoubleType(), True),
                                StructField('CMETH', StringType(), True),
                                StructField('CUOBF', DoubleType(), True),
                                StructField('KZUMW', StringType(), True),
                                StructField('KOSCH', StringType(), True),
                                StructField('SPROF', StringType(), True),
                                StructField('NRFHG', StringType(), True),
                                StructField('MFRPN', StringType(), True),
                                StructField('MFRNR', StringType(), True),
                                StructField('BMATN', StringType(), True),
                                StructField('MPROF', StringType(), True),
                                StructField('KZWSM', StringType(), True),
                                StructField('SAITY', StringType(), True),
                                StructField('PROFL', StringType(), True),
                                StructField('IHIVI', StringType(), True),
                                StructField('ILOOS', StringType(), True),
                                StructField('SERLV', StringType(), True),
                                StructField('KZGVH', StringType(), True),
                                StructField('XGCHP', StringType(), True),
                                StructField('KZEFF', StringType(), True),
                                StructField('COMPL', DoubleType(), True),
                                StructField('IPRKZ', StringType(), True),
                                StructField('RDMHD', StringType(), True),
                                StructField('PRZUS', StringType(), True),
                                StructField('MTPOS_MARA', StringType(), True),
                                StructField('BFLME', StringType(), True),
                                StructField('MATFI', StringType(), True),
                                StructField('CMREL', StringType(), True),
                                StructField('BBTYP', StringType(), True),
                                StructField('SLED_BBD', StringType(), True),
                                StructField('GTIN_VARIANT', StringType(), True),
                                StructField('GENNR', StringType(), True),
                                StructField('RMATP', StringType(), True),
                                StructField('GDS_RELEVANT', StringType(), True),
                                StructField('WEORA', StringType(), True),
                                StructField('HUTYP_DFLT', StringType(), True),
                                StructField('PILFERABLE', StringType(), True),
                                StructField('WHSTC', StringType(), True),
                                StructField('WHMATGR', StringType(), True),
                                StructField('HNDLCODE', StringType(), True),
                                StructField('HAZMAT', StringType(), True),
                                StructField('HUTYP', StringType(), True),
                                StructField('TARE_VAR', StringType(), True),
                                StructField('MAXC', DoubleType(), True),
                                StructField('MAXC_TOL', DoubleType(), True),
                                StructField('MAXL', DoubleType(), True),
                                StructField('MAXB', DoubleType(), True),
                                StructField('MAXH', DoubleType(), True),
                                StructField('MAXDIM_UOM', StringType(), True),
                                StructField('HERKL', StringType(), True),
                                StructField('MFRGR', StringType(), True),
                                StructField('QQTIME', DoubleType(), True),
                                StructField('QQTIMEUOM', StringType(), True),
                                StructField('QGRP', StringType(), True),
                                StructField('SERIAL', StringType(), True),
                                StructField('PS_SMARTFORM', StringType(), True),
                                StructField('LOGUNIT', StringType(), True),
                                StructField('CWQREL', StringType(), True),
                                StructField('CWQPROC', StringType(), True),
                                StructField('CWQTOLGR', StringType(), True),
                                StructField('/BEV1/LULEINH', DoubleType(), True),
                                StructField('/BEV1/LULDEGRP', StringType(), True),
                                StructField('/BEV1/NESTRUCCAT', StringType(), True),
                                StructField('/DSD/VC_GROUP', StringType(), True),
                                StructField('/VSO/R_TILT_IND', StringType(), True),
                                StructField('/VSO/R_STACK_IND', StringType(), True),
                                StructField('/VSO/R_BOT_IND', StringType(), True),
                                StructField('/VSO/R_TOP_IND', StringType(), True),
                                StructField('/VSO/R_STACK_NO', DoubleType(), True),
                                StructField('/VSO/R_PAL_IND', StringType(), True),
                                StructField('/VSO/R_PAL_OVR_D', DoubleType(), True),
                                StructField('/VSO/R_PAL_OVR_W', DoubleType(), True),
                                StructField('/VSO/R_PAL_B_HT', DoubleType(), True),
                                StructField('/VSO/R_PAL_MIN_H', DoubleType(), True),
                                StructField('/VSO/R_TOL_B_HT', DoubleType(), True),
                                StructField('/VSO/R_NO_P_GVH', DoubleType(), True),
                                StructField('/VSO/R_QUAN_UNIT', StringType(), True),
                                StructField('/VSO/R_KZGVH_IND', StringType(), True),
                                StructField('MCOND', StringType(), True),
                                StructField('RETDELC', StringType(), True),
                                StructField('LOGLEV_RETO', StringType(), True),
                                StructField('NSNID', StringType(), True),
                                StructField('IMATN', StringType(), True),
                                StructField('PICNUM', StringType(), True),
                                StructField('BSTAT', StringType(), True),
                                StructField('COLOR_ATINN', DoubleType(), True),
                                StructField('SIZE1_ATINN', DoubleType(), True),
                                StructField('SIZE2_ATINN', DoubleType(), True),
                                StructField('COLOR', StringType(), True),
                                StructField('SIZE1', StringType(), True),
                                StructField('SIZE2', StringType(), True),
                                StructField('FREE_CHAR', StringType(), True),
                                StructField('CARE_CODE', StringType(), True),
                                StructField('BRAND_ID', StringType(), True),
                                StructField('FIBER_CODE1', StringType(), True),
                                StructField('FIBER_PART1', DoubleType(), True),
                                StructField('FIBER_CODE2', StringType(), True),
                                StructField('FIBER_PART2', DoubleType(), True),
                                StructField('FIBER_CODE3', StringType(), True),
                                StructField('FIBER_PART3', DoubleType(), True),
                                StructField('FIBER_CODE4', StringType(), True),
                                StructField('FIBER_PART4', DoubleType(), True),
                                StructField('FIBER_CODE5', StringType(), True),
                                StructField('FIBER_PART5', DoubleType(), True),
                                StructField('FASHGRD', StringType(), True),
                                StructField('FECCARGA', TimestampType(), True),
                                StructField('FECULTACT', TimestampType(), True),
                                StructField('USRCARGA', StringType(), True)]),
              "TMP_IND_ECC_MARD" : StructType([StructField('MANDT', StringType(), True),
                                StructField('MATNR', StringType(), True),
                                StructField('WERKS', StringType(), True),
                                StructField('LGORT', StringType(), True),
                                StructField('PSTAT', StringType(), True),
                                StructField('LVORM', StringType(), True),
                                StructField('LFGJA', DoubleType(), True),
                                StructField('LFMON', DoubleType(), True),
                                StructField('SPERR', StringType(), True),
                                StructField('LABST', DoubleType(), True),
                                StructField('UMLME', DoubleType(), True),
                                StructField('INSME', DoubleType(), True),
                                StructField('EINME', DoubleType(), True),
                                StructField('SPEME', DoubleType(), True),
                                StructField('RETME', DoubleType(), True),
                                StructField('VMLAB', DoubleType(), True),
                                StructField('VMUML', DoubleType(), True),
                                StructField('VMINS', DoubleType(), True),
                                StructField('VMEIN', DoubleType(), True),
                                StructField('VMSPE', DoubleType(), True),
                                StructField('VMRET', DoubleType(), True),
                                StructField('KZILL', StringType(), True),
                                StructField('KZILQ', StringType(), True),
                                StructField('KZILE', StringType(), True),
                                StructField('KZILS', StringType(), True),
                                StructField('KZVLL', StringType(), True),
                                StructField('KZVLQ', StringType(), True),
                                StructField('KZVLE', StringType(), True),
                                StructField('KZVLS', StringType(), True),
                                StructField('DISKZ', StringType(), True),
                                StructField('LSOBS', StringType(), True),
                                StructField('LMINB', DoubleType(), True),
                                StructField('LBSTF', DoubleType(), True),
                                StructField('HERKL', StringType(), True),
                                StructField('EXPPG', StringType(), True),
                                StructField('EXVER', StringType(), True),
                                StructField('LGPBE', StringType(), True),
                                StructField('KLABS', DoubleType(), True),
                                StructField('KINSM', DoubleType(), True),
                                StructField('KEINM', DoubleType(), True),
                                StructField('KSPEM', DoubleType(), True),
                                StructField('DLINL', DateType(), True),
                                StructField('PRCTL', StringType(), True),
                                StructField('ERSDA', DateType(), True),
                                StructField('VKLAB', DoubleType(), True),
                                StructField('VKUML', DoubleType(), True),
                                StructField('LWMKB', StringType(), True),
                                StructField('BSKRF', DoubleType(), True),
                                StructField('MDRUE', StringType(), True),
                                StructField('MDJIN', DoubleType(), True),
                                StructField('FECCARGA', TimestampType(), True),
                                StructField('FECULTACT', TimestampType(), True),
                                StructField('USRCARGA', StringType(), True)])
    }
    if table == False:
        return schema_dict.keys()
    else: 
        return schema_dict[table]
        return schema_dict[table]

# COMMAND ----------

def get_schema_sap(table = False):
    schema_dict = {
        "MARA": StructType([
    StructField('MANDT', StringType(), True),
    StructField('MATNR', StringType(), True),
    StructField('ERSDA', StringType(), True),
    StructField('CREATED_AT_TIME', StringType(), True),
    StructField('ERNAM', StringType(), True),
    StructField('LAEDA', StringType(), True),
    StructField('AENAM', StringType(), True),
    StructField('VPSTA', StringType(), True),
    StructField('PSTAT', StringType(), True),
    StructField('LVORM', StringType(), True),
    StructField('MTART', StringType(), True),
    StructField('MBRSH', StringType(), True),
    StructField('MATKL', StringType(), True),
    StructField('BISMT', StringType(), True),
    StructField('MEINS', StringType(), True),
    StructField('BSTME', StringType(), True),
    StructField('ZEINR', StringType(), True),
    StructField('ZEIAR', StringType(), True),
    StructField('ZEIVR', StringType(), True),
    StructField('ZEIFO', StringType(), True),
    StructField('AESZN', StringType(), True),
    StructField('BLATT', StringType(), True),
    StructField('BLANZ', StringType(), True),
    StructField('FERTH', StringType(), True),
    StructField('FORMT', StringType(), True),
    StructField('GROES', StringType(), True),
    StructField('WRKST', StringType(), True),
    StructField('NORMT', StringType(), True),
    StructField('LABOR', StringType(), True),
    StructField('EKWSL', StringType(), True),
    StructField('BRGEW', StringType(), True),
    StructField('NTGEW', StringType(), True),
    StructField('GEWEI', StringType(), True),
    StructField('VOLUM', StringType(), True),
    StructField('VOLEH', StringType(), True),
    StructField('BEHVO', StringType(), True),
    StructField('RAUBE', StringType(), True),
    StructField('TEMPB', StringType(), True),
    StructField('DISST', StringType(), True),
    StructField('TRAGR', StringType(), True),
    StructField('STOFF', StringType(), True),
    StructField('SPART', StringType(), True),
    StructField('KUNNR', StringType(), True),
    StructField('EANNR', StringType(), True),
    StructField('WESCH', StringType(), True),
    StructField('BWVOR', StringType(), True),
    StructField('BWSCL', StringType(), True),
    StructField('SAISO', StringType(), True),
    StructField('ETIAR', StringType(), True),
    StructField('ETIFO', StringType(), True),
    StructField('ENTAR', StringType(), True),
    StructField('EAN11', StringType(), True),
    StructField('NUMTP', StringType(), True),
    StructField('LAENG', StringType(), True),
    StructField('BREIT', StringType(), True),
    StructField('HOEHE', StringType(), True),
    StructField('MEABM', StringType(), True),
    StructField('PRDHA', StringType(), True),
    StructField('AEKLK', StringType(), True),
    StructField('CADKZ', StringType(), True),
    StructField('QMPUR', StringType(), True),
    StructField('ERGEW', StringType(), True),
    StructField('ERGEI', StringType(), True),
    StructField('ERVOL', StringType(), True),
    StructField('ERVOE', StringType(), True),
    StructField('GEWTO', StringType(), True),
    StructField('VOLTO', StringType(), True),
    StructField('VABME', StringType(), True),
    StructField('KZREV', StringType(), True),
    StructField('KZKFG', StringType(), True),
    StructField('XCHPF', StringType(), True),
    StructField('VHART', StringType(), True),
    StructField('FUELG', StringType(), True),
    StructField('STFAK', StringType(), True),
    StructField('MAGRV', StringType(), True),
    StructField('BEGRU', StringType(), True),
    StructField('DATAB', StringType(), True),
    StructField('LIQDT', StringType(), True),
    StructField('SAISJ', StringType(), True),
    StructField('PLGTP', StringType(), True),
    StructField('MLGUT', StringType(), True),
    StructField('EXTWG', StringType(), True),
    StructField('SATNR', StringType(), True),
    StructField('ATTYP', StringType(), True),
    StructField('KZKUP', StringType(), True),
    StructField('KZNFM', StringType(), True),
    StructField('PMATA', StringType(), True),
    StructField('MSTAE', StringType(), True),
    StructField('MSTAV', StringType(), True),
    StructField('MSTDE', StringType(), True),
    StructField('MSTDV', StringType(), True),
    StructField('TAKLV', StringType(), True),
    StructField('RBNRM', StringType(), True),
    StructField('MHDRZ', StringType(), True),
    StructField('MHDHB', StringType(), True),
    StructField('MHDLP', StringType(), True),
    StructField('INHME', StringType(), True),
    StructField('INHAL', StringType(), True),
    StructField('VPREH', StringType(), True),
    StructField('ETIAG', StringType(), True),
    StructField('INHBR', StringType(), True),
    StructField('CMETH', StringType(), True),
    StructField('CUOBF', StringType(), True),
    StructField('KZUMW', StringType(), True),
    StructField('KOSCH', StringType(), True),
    StructField('SPROF', StringType(), True),
    StructField('NRFHG', StringType(), True),
    StructField('MFRPN', StringType(), True),
    StructField('MFRNR', StringType(), True),
    StructField('BMATN', StringType(), True),
    StructField('MPROF', StringType(), True),
    StructField('KZWSM', StringType(), True),
    StructField('SAITY', StringType(), True),
    StructField('PROFL', StringType(), True),
    StructField('IHIVI', StringType(), True),
    StructField('ILOOS', StringType(), True),
    StructField('SERLV', StringType(), True),
    StructField('KZGVH', StringType(), True),
    StructField('XGCHP', StringType(), True),
    StructField('KZEFF', StringType(), True),
    StructField('COMPL', StringType(), True),
    StructField('IPRKZ', StringType(), True),
    StructField('RDMHD', StringType(), True),
    StructField('PRZUS', StringType(), True),
    StructField('MTPOS_MARA', StringType(), True),
    StructField('BFLME', StringType(), True),
    StructField('MATFI', StringType(), True),
    StructField('CMREL', StringType(), True),
    StructField('BBTYP', StringType(), True),
    StructField('SLED_BBD', StringType(), True),
    StructField('GTIN_VARIANT', StringType(), True),
    StructField('GENNR', StringType(), True),
    StructField('RMATP', StringType(), True),
    StructField('GDS_RELEVANT', StringType(), True),
    StructField('WEORA', StringType(), True),
    StructField('HUTYP_DFLT', StringType(), True),
    StructField('PILFERABLE', StringType(), True),
    StructField('WHSTC', StringType(), True),
    StructField('WHMATGR', StringType(), True),
    StructField('HNDLCODE', StringType(), True),
    StructField('HAZMAT', StringType(), True),
    StructField('HUTYP', StringType(), True),
    StructField('TARE_VAR', StringType(), True),
    StructField('MAXC', StringType(), True),
    StructField('MAXC_TOL', StringType(), True),
    StructField('MAXL', StringType(), True),
    StructField('MAXB', StringType(), True),
    StructField('MAXH', StringType(), True),
    StructField('MAXDIM_UOM', StringType(), True),
    StructField('HERKL', StringType(), True),
    StructField('MFRGR', StringType(), True),
    StructField('QQTIME', StringType(), True),
    StructField('QQTIMEUOM', StringType(), True),
    StructField('QGRP', StringType(), True),
    StructField('SERIAL', StringType(), True),
    StructField('PS_SMARTFORM', StringType(), True),
    StructField('LOGUNIT', StringType(), True),
    StructField('CWQREL', StringType(), True),
    StructField('CWQPROC', StringType(), True),
    StructField('CWQTOLGR', StringType(), True),
    StructField('ADPROF', StringType(), True),
    StructField('IPMIPPRODUCT', StringType(), True),
    StructField('ALLOW_PMAT_IGNO', StringType(), True),
    StructField('MEDIUM', StringType(), True),
    StructField('COMMODITY', StringType(), True),
    StructField('ANIMAL_ORIGIN', StringType(), True),
    StructField('TEXTILE_COMP_IND', StringType(), True),
    StructField('LAST_CHANGED_TIME', StringType(), True),
    StructField('MATNR_EXTERNAL', StringType(), True),
    StructField('CHML_CMPLNC_RLVNCE_IND', StringType(), True),
    StructField('LOGISTICAL_MAT_CATEGORY', StringType(), True),
    StructField('SALES_MATERIAL', StringType(), True),
    StructField('IDENTIFICATION_TAG_TYPE', StringType(), True),
    StructField('SGT_CSGR', StringType(), True),
    StructField('SGT_COVSA', StringType(), True),
    StructField('SGT_STAT', StringType(), True),
    StructField('SGT_SCOPE', StringType(), True),
    StructField('SGT_REL', StringType(), True),
    StructField('ANP', StringType(), True),
    StructField('PSM_CODE', StringType(), True),
    StructField('FSH_MG_AT1', StringType(), True),
    StructField('FSH_MG_AT2', StringType(), True),
    StructField('FSH_MG_AT3', StringType(), True),
    StructField('FSH_SEALV', StringType(), True),
    StructField('FSH_SEAIM', StringType(), True),
    StructField('FSH_SC_MID', StringType(), True),
    StructField('DUMMY_PRD_INCL_EEW_PS', StringType(), True),
    StructField('SCM_MATID_GUID16', StringType(), True),
    StructField('SCM_MATID_GUID22', StringType(), True),
    StructField('SCM_MATURITY_DUR', StringType(), True),
    StructField('SCM_SHLF_LFE_REQ_MIN', StringType(), True),
    StructField('SCM_SHLF_LFE_REQ_MAX', StringType(), True),
    StructField('SCM_PUOM', StringType(), True),
    StructField('RMATP_PB', StringType(), True),
    StructField('PROD_SHAPE', StringType(), True),
    StructField('MO_PROFILE_ID', StringType(), True),
    StructField('OVERHANG_TRESH', StringType(), True),
    StructField('BRIDGE_TRESH', StringType(), True),
    StructField('BRIDGE_MAX_SLOPE', StringType(), True),
    StructField('HEIGHT_NONFLAT', StringType(), True),
    StructField('HEIGHT_NONFLAT_UOM', StringType(), True),
    StructField('SCM_KITCOMP', StringType(), True),
    StructField('SCM_PROD_PAOOPT', StringType(), True),
    StructField('SCM_BOD_DEPLVL', StringType(), True),
    StructField('SCM_RESTRICT_INVBAL', StringType(), True),
    StructField('SCM_DRP_GL_STOCK', StringType(), True),
    StructField('SCM_EXCL_EXPEDITE', StringType(), True),
    StructField('CWM_XCWMAT', StringType(), True),
    StructField('CWM_VALUM', StringType(), True),
    StructField('CWM_TOLGR', StringType(), True),
    StructField('CWM_TARA', StringType(), True),
    StructField('CWM_TARUM', StringType(), True),
    StructField('BEV1_LULEINH', StringType(), True),
    StructField('BEV1_LULDEGRP', StringType(), True),
    StructField('BEV1_NESTRUCCAT', StringType(), True),
    StructField('DSD_SL_TOLTYP', StringType(), True),
    StructField('DSD_SV_CNT_GRP', StringType(), True),
    StructField('DSD_VC_GROUP', StringType(), True),
    StructField('SAPMP_KADU', StringType(), True),
    StructField('SAPMP_ABMEIN', StringType(), True),
    StructField('SAPMP_KADP', StringType(), True),
    StructField('SAPMP_BRAD', StringType(), True),
    StructField('SAPMP_SPBI', StringType(), True),
    StructField('SAPMP_TRAD', StringType(), True),
    StructField('SAPMP_KEDU', StringType(), True),
    StructField('SAPMP_SPTR', StringType(), True),
    StructField('SAPMP_FBDK', StringType(), True),
    StructField('SAPMP_FBHK', StringType(), True),
    StructField('SAPMP_RILI', StringType(), True),
    StructField('SAPMP_FBAK', StringType(), True),
    StructField('SAPMP_AHO', StringType(), True),
    StructField('SAPMP_MIFRR', StringType(), True),
    StructField('STTPEC_SERTYPE', StringType(), True),
    StructField('STTPEC_SYNCACT', StringType(), True),
    StructField('STTPEC_SYNCTIME', StringType(), True),
    StructField('STTPEC_SYNCCHG', StringType(), True),
    StructField('STTPEC_COUNTRY_REF', StringType(), True),
    StructField('STTPEC_PRDCAT', StringType(), True),
    StructField('VSO_R_TILT_IND', StringType(), True),
    StructField('VSO_R_STACK_IND', StringType(), True),
    StructField('VSO_R_BOT_IND', StringType(), True),
    StructField('VSO_R_TOP_IND', StringType(), True),
    StructField('VSO_R_STACK_NO', StringType(), True),
    StructField('VSO_R_PAL_IND', StringType(), True),
    StructField('VSO_R_PAL_OVR_D', StringType(), True),
    StructField('VSO_R_PAL_OVR_W', StringType(), True),
    StructField('VSO_R_PAL_B_HT', StringType(), True),
    StructField('VSO_R_PAL_MIN_H', StringType(), True),
    StructField('VSO_R_TOL_B_HT', StringType(), True),
    StructField('VSO_R_NO_P_GVH', StringType(), True),
    StructField('VSO_R_QUAN_UNIT', StringType(), True),
    StructField('VSO_R_KZGVH_IND', StringType(), True),
    StructField('PACKCODE', StringType(), True),
    StructField('DG_PACK_STATUS', StringType(), True),
    StructField('SRV_DURA', StringType(), True),
    StructField('SRV_DURA_UOM', StringType(), True),
    StructField('SRV_SERWI', StringType(), True),
    StructField('SRV_ESCAL', StringType(), True),
    StructField('SOM_CYCLE', StringType(), True),
    StructField('SOM_CYCLE_RULE', StringType(), True),
    StructField('SOM_TC_SCHEMA', StringType(), True),
    StructField('SOM_CTR_AUTORENEWAL', StringType(), True),
    StructField('MCOND', StringType(), True),
    StructField('RETDELC', StringType(), True),
    StructField('LOGLEV_RETO', StringType(), True),
    StructField('NSNID', StringType(), True),
    StructField('ICFA', StringType(), True),
    StructField('RIC_ID', StringType(), True),
    StructField('DFS_SENSITIVITY_KEY', StringType(), True),
    StructField('DFS_MFRP2', StringType(), True),
    StructField('OVLPN', StringType(), True),
    StructField('ADSPC_SPC', StringType(), True),
    StructField('VARID', StringType(), True),
    StructField('MSBOOKPARTNO', StringType(), True),
    StructField('TOLERANCE_TYPE', StringType(), True),
    StructField('DPCBT', StringType(), True),
    StructField('XGRDT', StringType(), True),
    StructField('IMATN', StringType(), True),
    StructField('PICNUM', StringType(), True),
    StructField('BSTAT', StringType(), True),
    StructField('COLOR_ATINN', StringType(), True),
    StructField('SIZE1_ATINN', StringType(), True),
    StructField('SIZE2_ATINN', StringType(), True),
    StructField('COLOR', StringType(), True),
    StructField('SIZE1', StringType(), True),
    StructField('SIZE2', StringType(), True),
    StructField('FREE_CHAR', StringType(), True),
    StructField('CARE_CODE', StringType(), True),
    StructField('BRAND_ID', StringType(), True),
    StructField('FIBER_CODE1', StringType(), True),
    StructField('FIBER_PART1', StringType(), True),
    StructField('FIBER_CODE2', StringType(), True),
    StructField('FIBER_PART2', StringType(), True),
    StructField('FIBER_CODE3', StringType(), True),
    StructField('FIBER_PART3', StringType(), True),
    StructField('FIBER_CODE4', StringType(), True),
    StructField('FIBER_PART4', StringType(), True),
    StructField('FIBER_CODE5', StringType(), True),
    StructField('FIBER_PART5', StringType(), True),
    StructField('FASHGRD', StringType(), True)
]),
        
        "MBEW": StructType([StructField('MANDT', StringType(), True), StructField('MATNR', StringType(), True), StructField('BWKEY', StringType(), True), StructField('BWTAR', StringType(), True), StructField('LVORM', StringType(), True), StructField('LBKUM', StringType(), True), StructField('SALK3', StringType(), True), StructField('VPRSV', StringType(), True), StructField('VERPR', StringType(), True), StructField('STPRS', StringType(), True), StructField('PEINH', StringType(), True), StructField('BKLAS', StringType(), True), StructField('SALKV', StringType(), True), StructField('VMKUM', StringType(), True), StructField('VMSAL', StringType(), True), StructField('VMVPR', StringType(), True), StructField('VMVER', StringType(), True), StructField('VMSTP', StringType(), True), StructField('VMPEI', StringType(), True), StructField('VMBKL', StringType(), True), StructField('VMSAV', StringType(), True), StructField('VJKUM', StringType(), True), StructField('VJSAL', StringType(), True), StructField('VJVPR', StringType(), True), StructField('VJVER', StringType(), True), StructField('VJSTP', StringType(), True), StructField('VJPEI', StringType(), True), StructField('VJBKL', StringType(), True), StructField('VJSAV', StringType(), True), StructField('LFGJA', StringType(), True), StructField('LFMON', StringType(), True), StructField('BWTTY', StringType(), True), StructField('STPRV', StringType(), True), StructField('LAEPR', StringType(), True), StructField('ZKPRS', StringType(), True), StructField('ZKDAT', StringType(), True), StructField('TIMESTAMP', StringType(), True), StructField('BWPRS', StringType(), True), StructField('BWPRH', StringType(), True), StructField('VJBWS', StringType(), True), StructField('VJBWH', StringType(), True), StructField('VVJSL', StringType(), True), StructField('VVJLB', StringType(), True), StructField('VVMLB', StringType(), True), StructField('VVSAL', StringType(), True), StructField('ZPLPR', StringType(), True), StructField('ZPLP1', StringType(), True), StructField('ZPLP2', StringType(), True), StructField('ZPLP3', StringType(), True), StructField('ZPLD1', StringType(), True), StructField('ZPLD2', StringType(), True), StructField('ZPLD3', StringType(), True), StructField('PPERZ', StringType(), True), StructField('PPERL', StringType(), True), StructField('PPERV', StringType(), True), StructField('KALKZ', StringType(), True), StructField('KALKL', StringType(), True), StructField('KALKV', StringType(), True), StructField('KALSC', StringType(), True), StructField('XLIFO', StringType(), True), StructField('MYPOL', StringType(), True), StructField('BWPH1', StringType(), True), StructField('BWPS1', StringType(), True), StructField('ABWKZ', StringType(), True), StructField('PSTAT', StringType(), True), StructField('KALN1', StringType(), True), StructField('KALNR', StringType(), True), StructField('BWVA1', StringType(), True), StructField('BWVA2', StringType(), True), StructField('BWVA3', StringType(), True), StructField('VERS1', StringType(), True), StructField('VERS2', StringType(), True), StructField('VERS3', StringType(), True), StructField('HRKFT', StringType(), True), StructField('KOSGR', StringType(), True), StructField('PPRDZ', StringType(), True), StructField('PPRDL', StringType(), True), StructField('PPRDV', StringType(), True), StructField('PDATZ', StringType(), True), StructField('PDATL', StringType(), True), StructField('PDATV', StringType(), True), StructField('EKALR', StringType(), True), StructField('VPLPR', StringType(), True), StructField('MLMAA', StringType(), True), StructField('MLAST', StringType(), True), StructField('LPLPR', StringType(), True), StructField('VKSAL', StringType(), True), StructField('HKMAT', StringType(), True), StructField('SPERW', StringType(), True), StructField('KZIWL', StringType(), True), StructField('WLINL', StringType(), True), StructField('ABCIW', StringType(), True), StructField('BWSPA', StringType(), True), StructField('LPLPX', StringType(), True), StructField('VPLPX', StringType(), True), StructField('FPLPX', StringType(), True), StructField('LBWST', StringType(), True), StructField('VBWST', StringType(), True), StructField('FBWST', StringType(), True), StructField('EKLAS', StringType(), True), StructField('QKLAS', StringType(), True), StructField('MTUSE', StringType(), True), StructField('MTORG', StringType(), True), StructField('OWNPR', StringType(), True), StructField('XBEWM', StringType(), True), StructField('BWPEI', StringType(), True), StructField('MBRUE', StringType(), True), StructField('OKLAS', StringType(), True), StructField('DUMMY_VAL_INCL_EEW_PS', StringType(), True), StructField('OIPPINV', StringType(), True)]),
    
            "KONP": StructType([StructField('MANDT', StringType(), True),
                                StructField('KNUMH', StringType(), True),
                                StructField('KOPOS', StringType(), True),
                                StructField('KAPPL', StringType(), True),
                                StructField('KSCHL', StringType(), True),
                                StructField('STFKZ', StringType(), True),
                                StructField('KSTBM', StringType(), True),
                                StructField('KRECH', StringType(), True),
                                StructField('KBETR', StringType(), True),
                                StructField('KONWA', StringType(), True),
                                StructField('KONMS', StringType(), True),
                                StructField('KPEIN', StringType(), True),
                                StructField('KMEIN', StringType(), True),
                                StructField('KUMZA', StringType(), True),
                                StructField('KUMNE', StringType(), True),
                                StructField('MXWRT', StringType(), True),
                                StructField('GKWRT', StringType(), True),
                                StructField('PKWRT', StringType(), True),
                                StructField('FKWRT', StringType(), True),
                                StructField('RSWRT', StringType(), True),
                                StructField('UKBAS', StringType(), True),
                                StructField('KUNNR', StringType(), True),
                                StructField('ZAEHK_IND', StringType(), True),
                                StructField('KBRUE', StringType(), True)]),
    
    "MAKT": StructType([
        StructField('MANDT', StringType(), True),
        StructField('MATNR', StringType(), True),
        StructField('SPRAS', StringType(), True),
        StructField('MAKTX', StringType(), True),
        StructField('MAKTG', StringType(), True)
    ]),
    
    "MARC": StructType([
    StructField('MANDT', StringType(), True),
    StructField('MATNR', StringType(), True),
    StructField('WERKS', StringType(), True),
    StructField('PSTAT', StringType(), True),
    StructField('LVORM', StringType(), True),
    StructField('BWTTY', StringType(), True),
    StructField('XCHAR', StringType(), True),
    StructField('MMSTA', StringType(), True),
    StructField('MMSTD', StringType(), True),
    StructField('MAABC', StringType(), True),
    StructField('KZKRI', StringType(), True),
    StructField('EKGRP', StringType(), True),
    StructField('AUSME', StringType(), True),
    StructField('DISPR', StringType(), True),
    StructField('DISMM', StringType(), True),
    StructField('DISPO', StringType(), True),
    StructField('KZDIE', StringType(), True),
    StructField('PLIFZ', StringType(), True),
    StructField('WEBAZ', StringType(), True),
    StructField('PERKZ', StringType(), True),
    StructField('AUSSS', StringType(), True),
    StructField('DISLS', StringType(), True),
    StructField('BESKZ', StringType(), True),
    StructField('SOBSL', StringType(), True),
    StructField('MINBE', StringType(), True),
    StructField('EISBE', StringType(), True),
    StructField('BSTMI', StringType(), True),
    StructField('BSTMA', StringType(), True),
    StructField('BSTFE', StringType(), True),
    StructField('BSTRF', StringType(), True),
    StructField('MABST', StringType(), True),
    StructField('LOSFX', StringType(), True),
    StructField('SBDKZ', StringType(), True),
    StructField('LAGPR', StringType(), True),
    StructField('ALTSL', StringType(), True),
    StructField('KZAUS', StringType(), True),
    StructField('AUSDT', StringType(), True),
    StructField('NFMAT', StringType(), True),
    StructField('KZBED', StringType(), True),
    StructField('MISKZ', StringType(), True),
    StructField('FHORI', StringType(), True),
    StructField('PFREI', StringType(), True),
    StructField('FFREI', StringType(), True),
    StructField('RGEKZ', StringType(), True),
    StructField('FEVOR', StringType(), True),
    StructField('BEARZ', StringType(), True),
    StructField('RUEZT', StringType(), True),
    StructField('TRANZ', StringType(), True),
    StructField('BASMG', StringType(), True),
    StructField('DZEIT', StringType(), True),
    StructField('MAXLZ', StringType(), True),
    StructField('LZEIH', StringType(), True),
    StructField('KZPRO', StringType(), True),
    StructField('GPMKZ', StringType(), True),
    StructField('UEETO', StringType(), True),
    StructField('UEETK', StringType(), True),
    StructField('UNETO', StringType(), True),
    StructField('WZEIT', StringType(), True),
    StructField('ATPKZ', StringType(), True),
    StructField('VZUSL', StringType(), True),
    StructField('HERBL', StringType(), True),
    StructField('INSMK', StringType(), True),
    StructField('SPROZ', StringType(), True),
    StructField('QUAZT', StringType(), True),
    StructField('SSQSS', StringType(), True),
    StructField('MPDAU', StringType(), True),
    StructField('KZPPV', StringType(), True),
    StructField('KZDKZ', StringType(), True),
    StructField('WSTGH', StringType(), True),
    StructField('PRFRQ', StringType(), True),
    StructField('NKMPR', StringType(), True),
    StructField('UMLMC', StringType(), True),
    StructField('LADGR', StringType(), True),
    StructField('XCHPF', StringType(), True),
    StructField('USEQU', StringType(), True),
    StructField('LGRAD', StringType(), True),
    StructField('AUFTL', StringType(), True),
    StructField('PLVAR', StringType(), True),
    StructField('OTYPE', StringType(), True),
    StructField('OBJID', StringType(), True),
    StructField('MTVFP', StringType(), True),
    StructField('PERIV', StringType(), True),
    StructField('KZKFK', StringType(), True),
    StructField('VRVEZ', StringType(), True),
    StructField('VBAMG', StringType(), True),
    StructField('VBEAZ', StringType(), True),
    StructField('LIZYK', StringType(), True),
    StructField('BWSCL', StringType(), True),
    StructField('KAUTB', StringType(), True),
    StructField('KORDB', StringType(), True),
    StructField('STAWN', StringType(), True),
    StructField('HERKL', StringType(), True),
    StructField('HERKR', StringType(), True),
    StructField('EXPME', StringType(), True),
    StructField('MTVER', StringType(), True),
    StructField('PRCTR', StringType(), True),
    StructField('TRAME', StringType(), True),
    StructField('MRPPP', StringType(), True),
    StructField('SAUFT', StringType(), True),
    StructField('FXHOR', StringType(), True),
    StructField('VRMOD', StringType(), True),
    StructField('VINT1', StringType(), True),
    StructField('VINT2', StringType(), True),
    StructField('VERKZ', StringType(), True),
    StructField('STLAL', StringType(), True),
    StructField('STLAN', StringType(), True),
    StructField('PLNNR', StringType(), True),
    StructField('APLAL', StringType(), True),
    StructField('LOSGR', StringType(), True),
    StructField('SOBSK', StringType(), True),
    StructField('FRTME', StringType(), True),
    StructField('LGPRO', StringType(), True),
    StructField('DISGR', StringType(), True),
    StructField('KAUSF', StringType(), True),
    StructField('QZGTP', StringType(), True),
    StructField('QMATV', StringType(), True),
    StructField('TAKZT', StringType(), True),
    StructField('RWPRO', StringType(), True),
    StructField('COPAM', StringType(), True),
    StructField('ABCIN', StringType(), True),
    StructField('AWSLS', StringType(), True),
    StructField('SERNP', StringType(), True),
    StructField('CUOBJ', StringType(), True),
    StructField('STDPD', StringType(), True),
    StructField('SFEPR', StringType(), True),
    StructField('XMCNG', StringType(), True),
    StructField('QSSYS', StringType(), True),
    StructField('LFRHY', StringType(), True),
    StructField('RDPRF', StringType(), True),
    StructField('VRBMT', StringType(), True),
    StructField('VRBWK', StringType(), True),
    StructField('VRBDT', StringType(), True),
    StructField('VRBFK', StringType(), True),
    StructField('AUTRU', StringType(), True),
    StructField('PREFE', StringType(), True),
    StructField('PRENC', StringType(), True),
    StructField('PRENO', StringType(), True),
    StructField('PREND', StringType(), True),
    StructField('PRENE', StringType(), True),
    StructField('PRENG', StringType(), True),
    StructField('ITARK', StringType(), True),
    StructField('SERVG', StringType(), True),
    StructField('KZKUP', StringType(), True),
    StructField('STRGR', StringType(), True),
    StructField('CUOBV', StringType(), True),
    StructField('LGFSB', StringType(), True),
    StructField('SCHGT', StringType(), True),
    StructField('CCFIX', StringType(), True),
    StructField('EPRIO', StringType(), True),
    StructField('QMATA', StringType(), True),
    StructField('RESVP', StringType(), True),
    StructField('PLNTY', StringType(), True),
    StructField('UOMGR', StringType(), True),
    StructField('UMRSL', StringType(), True),
    StructField('ABFAC', StringType(), True),
    StructField('SFCPF', StringType(), True),
    StructField('SHFLG', StringType(), True),
    StructField('SHZET', StringType(), True),
    StructField('MDACH', StringType(), True),
    StructField('KZECH', StringType(), True),
    StructField('MEGRU', StringType(), True),
    StructField('MFRGR', StringType(), True),
    StructField('SFTY_STK_METH', StringType(), True),
    StructField('PROFIL', StringType(), True),
    StructField('VKUMC', StringType(), True),
    StructField('VKTRW', StringType(), True),
    StructField('KZAGL', StringType(), True),
    StructField('FVIDK', StringType(), True),
    StructField('FXPRU', StringType(), True),
    StructField('LOGGR', StringType(), True),
    StructField('FPRFM', StringType(), True),
    StructField('GLGMG', StringType(), True),
    StructField('VKGLG', StringType(), True),
    StructField('INDUS', StringType(), True),
    StructField('MOWNR', StringType(), True),
    StructField('MOGRU', StringType(), True),
    StructField('CASNR', StringType(), True),
    StructField('GPNUM', StringType(), True),
    StructField('STEUC', StringType(), True),
    StructField('FABKZ', StringType(), True),
    StructField('MATGR', StringType(), True),
    StructField('VSPVB', StringType(), True),
    StructField('DPLFS', StringType(), True),
    StructField('DPLPU', StringType(), True),
    StructField('DPLHO', StringType(), True),
    StructField('MINLS', StringType(), True),
    StructField('MAXLS', StringType(), True),
    StructField('FIXLS', StringType(), True),
    StructField('LTINC', StringType(), True),
    StructField('COMPL', StringType(), True),
    StructField('CONVT', StringType(), True),
    StructField('SHPRO', StringType(), True),
    StructField('AHDIS', StringType(), True),
    StructField('DIBER', StringType(), True),
    StructField('KZPSP', StringType(), True),
    StructField('OCMPF', StringType(), True),
    StructField('APOKZ', StringType(), True),
    StructField('MCRUE', StringType(), True),
    StructField('LFMON', StringType(), True),
    StructField('LFGJA', StringType(), True),
    StructField('EISLO', StringType(), True),
    StructField('NCOST', StringType(), True),
    StructField('ROTATION_DATE', StringType(), True),
    StructField('UCHKZ', StringType(), True),
    StructField('UCMAT', StringType(), True),
    StructField('EXCISE_TAX_RLVNCE', StringType(), True),
    StructField('BWESB', StringType(), True),
    StructField('SGT_COVS', StringType(), True),
    StructField('SGT_STATC', StringType(), True),
    StructField('SGT_SCOPE', StringType(), True),
    StructField('SGT_MRPSI', StringType(), True),
    StructField('SGT_PRCM', StringType(), True),
    StructField('SGT_CHINT', StringType(), True),
    StructField('SGT_STK_PRT', StringType(), True),
    StructField('SGT_DEFSC', StringType(), True),
    StructField('SGT_MRP_ATP_STATUS', StringType(), True),
    StructField('SGT_MMSTD', StringType(), True),
    StructField('FSH_MG_ARUN_REQ', StringType(), True),
    StructField('FSH_SEAIM', StringType(), True),
    StructField('FSH_VAR_GROUP', StringType(), True),
    StructField('FSH_KZECH', StringType(), True),
    StructField('FSH_CALENDAR_GROUP', StringType(), True),
    StructField('ARUN_FIX_BATCH', StringType(), True),
    StructField('PPSKZ', StringType(), True),
    StructField('CONS_PROCG', StringType(), True),
    StructField('GI_PR_TIME', StringType(), True),
    StructField('MULTIPLE_EKGRP', StringType(), True),
    StructField('REF_SCHEMA', StringType(), True),
    StructField('MIN_TROC', StringType(), True),
    StructField('MAX_TROC', StringType(), True),
    StructField('TARGET_STOCK', StringType(), True),
    StructField('NF_FLAG', StringType(), True),
    StructField('CWM_UMLMC', StringType(), True),
    StructField('CWM_TRAME', StringType(), True),
    StructField('CWM_BWESB', StringType(), True),
    StructField('SCM_MATLOCID_GUID16', StringType(), True),
    StructField('SCM_MATLOCID_GUID22', StringType(), True),
    StructField('SCM_GRPRT', StringType(), True),
    StructField('SCM_GIPRT', StringType(), True),
    StructField('SCM_SCOST', StringType(), True),
    StructField('SCM_RELDT', StringType(), True),
    StructField('SCM_RRP_TYPE', StringType(), True),
    StructField('SCM_HEUR_ID', StringType(), True),
    StructField('SCM_PACKAGE_ID', StringType(), True),
    StructField('SCM_SSPEN', StringType(), True),
    StructField('SCM_GET_ALERTS', StringType(), True),
    StructField('SCM_RES_NET_NAME', StringType(), True),
    StructField('SCM_CONHAP', StringType(), True),
    StructField('SCM_HUNIT', StringType(), True),
    StructField('SCM_CONHAP_OUT', StringType(), True),
    StructField('SCM_HUNIT_OUT', StringType(), True),
    StructField('SCM_SHELF_LIFE_LOC', StringType(), True),
    StructField('SCM_SHELF_LIFE_DUR', StringType(), True),
    StructField('SCM_MATURITY_DUR', StringType(), True),
    StructField('SCM_SHLF_LFE_REQ_MIN', StringType(), True),
    StructField('SCM_SHLF_LFE_REQ_MAX', StringType(), True),
    StructField('SCM_LSUOM', StringType(), True),
    StructField('SCM_REORD_DUR', StringType(), True),
    StructField('SCM_TARGET_DUR', StringType(), True),
    StructField('SCM_TSTRID', StringType(), True),
    StructField('SCM_STRA1', StringType(), True),
    StructField('SCM_PEG_PAST_ALERT', StringType(), True),
    StructField('SCM_PEG_FUTURE_ALERT', StringType(), True),
    StructField('SCM_PEG_STRATEGY', StringType(), True),
    StructField('SCM_PEG_WO_ALERT_FST', StringType(), True),
    StructField('SCM_FIXPEG_PROD_SET', StringType(), True),
    StructField('SCM_WHATBOM', StringType(), True),
    StructField('SCM_RRP_SEL_GROUP', StringType(), True),
    StructField('SCM_INTSRC_PROF', StringType(), True),
    StructField('SCM_PRIO', StringType(), True),
    StructField('SCM_MIN_PASS_AMOUNT', StringType(), True),
    StructField('SCM_PROFID', StringType(), True),
    StructField('SCM_GES_MNG_USE', StringType(), True),
    StructField('SCM_GES_BST_USE', StringType(), True),
    StructField('ESPPFLG', StringType(), True),
    StructField('SCM_THRUPUT_TIME', StringType(), True),
    StructField('SCM_TPOP', StringType(), True),
    StructField('SCM_SAFTY_V', StringType(), True),
    StructField('SCM_PPSAFTYSTK', StringType(), True),
    StructField('SCM_PPSAFTYSTK_V', StringType(), True),
    StructField('SCM_REPSAFTY', StringType(), True),
    StructField('SCM_REPSAFTY_V', StringType(), True),
    StructField('SCM_REORD_V', StringType(), True),
    StructField('SCM_MAXSTOCK_V', StringType(), True),
    StructField('SCM_SCOST_PRCNT', StringType(), True),
    StructField('SCM_PROC_COST', StringType(), True),
    StructField('SCM_NDCOSTWE', StringType(), True),
    StructField('SCM_NDCOSTWA', StringType(), True),
    StructField('SCM_CONINP', StringType(), True),
    StructField('CONF_GMSYNC', StringType(), True),
    StructField('SCM_IUNIT', StringType(), True),
    StructField('SCM_SFT_LOCK', StringType(), True),
    StructField('DUMMY_PLNT_INCL_EEW_PS', StringType(), True),
    StructField('SAPMP_TOLPRPL', StringType(), True),
    StructField('SAPMP_TOLPRMI', StringType(), True),
    StructField('STTPEC_SERVALID', StringType(), True),
    StructField('VSO_R_PKGRP', StringType(), True),
    StructField('VSO_R_LANE_NUM', StringType(), True),
    StructField('VSO_R_PAL_VEND', StringType(), True),
    StructField('VSO_R_FORK_DIR', StringType(), True),
    StructField('IUID_RELEVANT', StringType(), True),
    StructField('IUID_TYPE', StringType(), True),
    StructField('UID_IEA', StringType(), True),
    StructField('DPCBT', StringType(), True)
]),
        
        "MLGN": StructType([StructField('MANDT', StringType(), True), StructField('MATNR', StringType(), True), 
                            StructField('LGNUM', StringType(), True), StructField('LGBKZ', StringType(), True), StructField('LTKZE', StringType(), True), StructField('LTKZA', StringType(), True), StructField('LHMG1', StringType(), True), StructField('LHMG2', StringType(), True), StructField('LHMG3', StringType(), True), StructField('LHME1', StringType(), True), StructField('LHME2', StringType(), True), StructField('LETY1', StringType(), True), StructField('LETY2', StringType(), True), StructField('LETY3', StringType(), True), StructField('LVSME', StringType(), True), StructField('BLOCK', StringType(), True), StructField('BSSKZ', StringType(), True), StructField('MKAPV', StringType(), True), StructField('BEZME', StringType(), True), StructField('PLKPT', StringType(), True)]),
        
        "VBAP": StructType([StructField('MANDT', StringType(), True), StructField('VBELN', StringType(), True), 
                            StructField('POSNR', StringType(), True), StructField('MATNR', StringType(), True), StructField('MATWA', StringType(), True), StructField('MATKL', StringType(), True), StructField('ARKTX', StringType(), True), StructField('PSTYV', StringType(), True), StructField('FKREL', StringType(), True), StructField('ABGRU', StringType(), True), StructField('PRODH', StringType(), True), StructField('ZWERT', StringType(), True), StructField('ZMENG', StringType(), True), StructField('ZIEME', StringType(), True), StructField('UMZIZ', StringType(), True), StructField('UMZIN', StringType(), True), StructField('MEINS', StringType(), True), StructField('SUBSTN_NUMERATOR', StringType(), True), StructField('SUBSTN_DENOMINATOR', StringType(), True), StructField('SMENG', StringType(), True), StructField('ABLFZ', StringType(), True), StructField('ABSFZ', StringType(), True), StructField('KBVER', StringType(), True), StructField('KEVER', StringType(), True), StructField('UEBTO', StringType(), True), StructField('UNTTO', StringType(), True), StructField('SPART', StringType(), True), StructField('NETWR', StringType(), True), StructField('WAERK', StringType(), True), StructField('ANTLF', StringType(), True), StructField('KWMENG', StringType(), True), StructField('LSMENG', StringType(), True), StructField('KBMENG', StringType(), True), StructField('KLMENG', StringType(), True), StructField('VRKME', StringType(), True), StructField('UMVKZ', StringType(), True), StructField('UMVKN', StringType(), True), StructField('BRGEW', StringType(), True), StructField('NTGEW', StringType(), True), StructField('GEWEI', StringType(), True), StructField('VOLUM', StringType(), True), StructField('VOLEH', StringType(), True), StructField('WERKS', StringType(), True), StructField('LGORT', StringType(), True), StructField('VSTEL', StringType(), True), StructField('ROUTE', StringType(), True), StructField('STPOS', StringType(), True), StructField('AWAHR', StringType(), True), StructField('ERDAT', StringType(), True), StructField('ERNAM', StringType(), True), StructField('ERZET', StringType(), True), StructField('TAXM1', StringType(), True), StructField('VBEAF', StringType(), True), StructField('VBEAV', StringType(), True), StructField('NETPR', StringType(), True), StructField('KPEIN', StringType(), True), StructField('KMEIN', StringType(), True), StructField('SKTOF', StringType(), True), StructField('MTVFP', StringType(), True), StructField('SUMBD', StringType(), True), StructField('KONDM', StringType(), True), StructField('KTGRM', StringType(), True), StructField('PRSOK', StringType(), True), StructField('XCHPF', StringType(), True), StructField('XCHAR', StringType(), True), StructField('LFMNG', StringType(), True), StructField('STAFO', StringType(), True), StructField('WAVWR', StringType(), True), StructField('KZWI1', StringType(), True), StructField('KZWI2', StringType(), True), StructField('KZWI3', StringType(), True), StructField('KZWI4', StringType(), True), StructField('KZWI5', StringType(), True), StructField('KZWI6', StringType(), True), StructField('STCUR', StringType(), True), StructField('EAN11', StringType(), True), StructField('PRCTR', StringType(), True), StructField('MVGR1', StringType(), True), StructField('MVGR2', StringType(), True), StructField('MVGR3', StringType(), True), StructField('MVGR4', StringType(), True), StructField('MVGR5', StringType(), True), StructField('KMPMG', StringType(), True), StructField('SOBKZ', StringType(), True), StructField('BOB_FG_ID', StringType(), True), StructField('BOB_PROMOTION_ID', StringType(), True), StructField('UMREF', StringType(), True), StructField('BEDAE', StringType(), True), StructField('CMPRE', StringType(), True), StructField('CMPNT', StringType(), True), StructField('CMKUA', StringType(), True), StructField('ANZSN', StringType(), True), StructField('MAGRV', StringType(), True), StructField('CMPRE_FLT', StringType(), True), StructField('ABGES', StringType(), True), StructField('MWSBP', StringType(), True), StructField('BERID', StringType(), True), StructField('LOGSYS_EXT', StringType(), True), StructField('HANDOVERTIME', StringType(), True), StructField('ABSTA', StringType(), True), StructField('BESTA', StringType(), True), StructField('GBSTA', StringType(), True), StructField('LFGSA', StringType(), True), StructField('LFSTA', StringType(), True), StructField('UVALL', StringType(), True), StructField('UVFAK', StringType(), True), StructField('UVPRS', StringType(), True), StructField('UVVLK', StringType(), True), StructField('CMTD_DELIV_DATE', StringType(), True), StructField('CMTD_DELIV_CREADATE', StringType(), True), StructField('CMTD_DELIV_QTY_SU', StringType(), True), StructField('REQQTY_BU', StringType(), True), StructField('HANDLE', StringType(), True), StructField('IFRS15_TOTAL_SSP', StringType(), True), StructField('CAPPED_NET_AMOUNT', StringType(), True), StructField('SESSION_CREATION_DATE', StringType(), True), StructField('SESSION_CREATION_TIME', StringType(), True), StructField('VBTYP_ANA', StringType(), True), StructField('AUART_ANA', StringType(), True), StructField('VKORG_ANA', StringType(), True), StructField('VTWEG_ANA', StringType(), True), StructField('SPART_ANA', StringType(), True), StructField('VKBUR_ANA', StringType(), True), StructField('VKGRP_ANA', StringType(), True), StructField('AUDAT_ANA', StringType(), True), StructField('KVGR2_ANA', StringType(), True), StructField('KVGR3_ANA', StringType(), True), StructField('KVGR4_ANA', StringType(), True), StructField('VDATU_ANA', StringType(), True), StructField('VSBED_ANA', StringType(), True), StructField('KUNNR_ANA', StringType(), True), StructField('KNUMV_ANA', StringType(), True), StructField('BZIRK_ANA', StringType(), True), StructField('BSTKD_ANA', StringType(), True), StructField('FKDAT_ANA', StringType(), True), StructField('KUNWE_ANA', StringType(), True), StructField('KUNRE_ANA', StringType(), True), StructField('KUNRG_ANA', StringType(), True), StructField('CPD_UPDAT', StringType(), True), StructField('PRS_WORK_PERIOD', StringType(), True)
                            ]),
        
        "A004": StructType([StructField('MANDT', StringType(), True), StructField('MATNR', StringType(), True),
                             StructField('KAPPL', StringType(), True), StructField('KSCHL', StringType(), True), StructField('VKORG', StringType(), True), StructField('VTWEG', StringType(), True), StructField('DATBI', StringType(), True), StructField('DATAB', StringType(), True), StructField('KNUMH', StringType(), True)]),
        
        "ADRC": StructType([StructField('CLIENT', StringType(), True), StructField('ADDRNUMBER', StringType(), True),
                             StructField('DATE_FROM', StringType(), True), StructField('NATION', StringType(), True), StructField('DATE_TO', StringType(), True), StructField('TITLE', StringType(), True), StructField('NAME1', StringType(), True), StructField('CITY_CODE', StringType(), True), StructField('CITYP_CODE', StringType(), True), StructField('CITYH_CODE', StringType(), True), StructField('COUNTRY', StringType(), True), StructField('LANGU', StringType(), True), StructField('REGION', StringType(), True), StructField('ADDR_GROUP', StringType(), True), StructField('MC_NAME1', StringType(), True), StructField('TIME_ZONE', StringType(), True), StructField('LANGU_CREA', StringType(), True), StructField('ADRC_UUID', StringType(), True)]),
    
        
        "KNA1": StructType([StructField('MANDT', StringType(), True), StructField('KUNNR', StringType(), True),
                             StructField('LAND1', StringType(), True), StructField('NAME1', StringType(), True), StructField('NAME2', StringType(), True), StructField('ORT01', StringType(), True), StructField('PSTLZ', StringType(), True), StructField('REGIO', StringType(), True), StructField('SORTL', StringType(), True), StructField('STRAS', StringType(), True), StructField('TELF1', StringType(), True), StructField('TELFX', StringType(), True), StructField('XCPDK', StringType(), True), StructField('ADRNR', StringType(), True), StructField('MCOD1', StringType(), True), StructField('MCOD2', StringType(), True), StructField('MCOD3', StringType(), True), StructField('AUFSD', StringType(), True), StructField('ERDAT', StringType(), True), StructField('ERNAM', StringType(), True), StructField('FAKSD', StringType(), True), StructField('KTOKD', StringType(), True), StructField('LIFSD', StringType(), True), StructField('LOEVM', StringType(), True), StructField('NAME4', StringType(), True), StructField('ORT02', StringType(), True), StructField('SPRAS', StringType(), True), StructField('STCD1', StringType(), True), StructField('LZONE', StringType(), True), StructField('KATR7', StringType(), True), StructField('KATR10', StringType(), True), StructField('DUEFL', StringType(), True), StructField('CASSD', StringType(), True), StructField('UPTIM', StringType(), True), StructField('NODEL', StringType(), True), StructField('J_3GSTDMON', StringType(), True), StructField('J_3GSTDTAG', StringType(), True), StructField('J_3GTAGMON', StringType(), True)]),
        
        "MVKE": StructType([StructField('MANDT', StringType(), True), StructField('MATNR', StringType(), True), 
                            StructField('VKORG', StringType(), True), StructField('VTWEG', StringType(), True), StructField('SKTOF', StringType(), True), StructField('VMSTD', StringType(), True), StructField('AUMNG', StringType(), True), StructField('LFMNG', StringType(), True), StructField('EFMNG', StringType(), True), StructField('SCMNG', StringType(), True), StructField('MTPOS', StringType(), True), StructField('DWERK', StringType(), True), StructField('PRODH', StringType(), True), StructField('PROVG', StringType(), True), StructField('KONDM', StringType(), True), StructField('KTGRM', StringType(), True), StructField('MVGR1', StringType(), True), StructField('MVGR2', StringType(), True), StructField('MVGR3', StringType(), True), StructField('MVGR4', StringType(), True), StructField('MVGR5', StringType(), True), StructField('LFMAX', StringType(), True), StructField('VMSTA', StringType(), True), StructField('RDPRF', StringType(), True)]),
        
        # "VBBE": StructType([StructField('MANDT', StringType(), True), StructField('VBELN', StringType(), True), 
        #                     StructField('POSNR', StringType(), True), StructField('ETENR', StringType(), True), 
        #                     StructField('MATNR', StringType(), True), StructField('WERKS', StringType(), True), StructField('BERID', StringType(), True), StructField('MBDAT', StringType(), True), StructField('LGORT', StringType(), True), StructField('VBTYP', StringType(), True), StructField('BDART', StringType(), True), StructField('PLART', StringType(), True), StructField('OMENG', StringType(), True), StructField('VMENG', StringType(), True), StructField('MEINS', StringType(), True), StructField('AWAHR', StringType(), True), StructField('AUART', StringType(), True), StructField('KUNNR', StringType(), True), StructField('UMREF', StringType(), True), StructField('SOBKZ', StringType(), True), StructField('MONKZ', StringType(), True), StructField('FSH_RALLOC_QTY', StringType(), True)]),
                                
        "VBKD": StructType([StructField('MANDT', StringType(), True), 
                            StructField('VBELN', StringType(), True), StructField('POSNR', StringType(), True), StructField('KONDA', StringType(), True), StructField('KDGRP', StringType(), True), StructField('BZIRK', StringType(), True), StructField('INCO1', StringType(), True), StructField('INCO2', StringType(), True), StructField('KZAZU', StringType(), True), StructField('ZTERM', StringType(), True), StructField('KTGRD', StringType(), True), StructField('KURSK', StringType(), True), StructField('PRSDT', StringType(), True), StructField('FKDAT', StringType(), True), StructField('STCUR', StringType(), True), StructField('BSTDK', StringType(), True), StructField('BSARK', StringType(), True), StructField('KURSK_DAT', StringType(), True), StructField('KURRF_DAT', StringType(), True), StructField('AKPRZ', StringType(), True), StructField('BSTKD_M', StringType(), True), StructField('CAMPAIGN', StringType(), True), StructField('INCO2_L', StringType(), True)]),
        
        #La tabla en SAP se llama PRCD_ELEMENTS
        "PRDC_ELEMENTS": StructType([StructField('CLIENT', StringType(), True), StructField('KNUMV', StringType(), True),
                                      StructField('KPOSN', StringType(), True), StructField('STUNR', StringType(), True), StructField('ZAEHK', StringType(), True), StructField('KAPPL', StringType(), True), StructField('KSCHL', StringType(), True), StructField('KDATU', StringType(), True), StructField('KRECH', StringType(), True), StructField('KAWRT', StringType(), True), StructField('KBETR', StringType(), True), StructField('KPEIN', StringType(), True), StructField('KUMZA', StringType(), True), StructField('KUMNE', StringType(), True), StructField('KNTYP', StringType(), True), StructField('KSTAT', StringType(), True), StructField('KHERK', StringType(), True), StructField('KDIFF', StringType(), True), StructField('KWERT', StringType(), True), StructField('WAERK', StringType(), True), StructField('KSTEU', StringType(), True), StructField('KOAID', StringType(), True), StructField('KFAKTOR', StringType(), True), StructField('KFAKTOR1', StringType(), True), StructField('KSTBS', StringType(), True), StructField('KWERT_K', StringType(), True), StructField('KBFLAG', StringType(), True), StructField('KAWRT_K', StringType(), True), StructField('KAQTY', StringType(), True)]),
        
        "VBAK" : StructType([StructField('MANDT', StringType(), True), StructField('VBELN', StringType(), True), 
                             StructField('ERDAT', StringType(), True), StructField('ERZET', StringType(), True), StructField('ERNAM', StringType(), True), StructField('AUDAT', StringType(), True), StructField('VBTYP', StringType(), True), StructField('TRVOG', StringType(), True), 
                             StructField('AUART', StringType(), True), StructField('AUGRU', StringType(), True), StructField('NETWR', StringType(), True), StructField('WAERK', StringType(), True), StructField('VKORG', StringType(), True), StructField('VTWEG', StringType(), True), StructField('SPART', StringType(), True), StructField('VKGRP', StringType(), True), StructField('VKBUR', StringType(), True), StructField('KNUMV', StringType(), True), StructField('VDATU', StringType(), True), StructField('VPRGR', StringType(), True), 
                             StructField('KALSM', StringType(), True), StructField('VSBED', StringType(), True), StructField('FKARA', StringType(), True), StructField('AWAHR', StringType(), True), 
                             StructField('BSTNK', StringType(), True), StructField('IHREZ', StringType(), True), StructField('TELF1', StringType(), True), StructField('MAHZA', StringType(), True), StructField('KUNNR', StringType(), True), StructField('KVGR2', StringType(), True), StructField('KVGR3', StringType(), True), StructField('KVGR4', StringType(), True), StructField('KOKRS', StringType(), True), StructField('KKBER', StringType(), True), StructField('KNKLI', StringType(), True), StructField('CTLPC', StringType(), True), StructField('CMWAE', StringType(), True), StructField('BUKRS_VF', StringType(), True), StructField('ABHOV', StringType(), True), StructField('ABHOB', StringType(), True), StructField('VZEIT', StringType(), True), StructField('HANDLE', StringType(), True), StructField('UPD_TMSTMP', StringType(), True), 
                             StructField('LAST_CHANGED_BY_USER', StringType(), True), 
                             StructField('EXT_REV_TMSTMP', StringType(), True), StructField('CROSSITEM_PRC_DATE', StringType(), True), StructField('ABSTK', StringType(), True), StructField('FKSAK', StringType(), True), StructField('GBSTK', StringType(), True), StructField('UVALL', StringType(), True), StructField('UVALS', StringType(), True), StructField('UVFAK', StringType(), True), StructField('UVFAS', StringType(), True), StructField('UVPRS', StringType(), True), StructField('UVVLK', StringType(), True), StructField('UVVLS', StringType(), True), StructField('SDM_VERSION', StringType(), True), StructField('ZZTIPORELACION', StringType(), True)]),

    "A005": StructType([StructField('MANDT', StringType(), True), StructField('KAPPL', StringType(), True), 
                        StructField('KSCHL', StringType(), True), StructField('VKORG', StringType(), True), 
                        StructField('VTWEG', StringType(), True), StructField('KUNNR', StringType(), True), 
                        StructField('MATNR', StringType(), True), StructField('DATBI', StringType(), True), 
                        StructField('DATAB', StringType(), True), StructField('KNUMH', StringType(), True)]),
    
    "VBFA" : StructType([StructField('MANDT', StringType(), True), 
                         StructField('RUUID', StringType(), True), StructField('VBELV', StringType(), True), StructField('POSNV', StringType(), True), StructField('VBELN', StringType(), True), StructField('POSNN', StringType(), True), StructField('VBTYP_N', StringType(), True), StructField('RFMNG', StringType(), True), StructField('MEINS', StringType(), True), StructField('RFWRT', StringType(), True), StructField('VBTYP_V', StringType(), True), StructField('PLMIN', StringType(), True), StructField('ERDAT', StringType(), True), StructField('ERZET', StringType(), True), StructField('MATNR', StringType(), True), StructField('RFMNG_FLO', StringType(), True), StructField('RFMNG_FLT', StringType(), True), StructField('VRKME', StringType(), True), StructField('ABGES', StringType(), True), StructField('SOBKZ', StringType(), True)]),
    
    "VBPA":StructType([StructField('MANDT', StringType(), True), StructField('VBELN', StringType(), True), 
                       StructField('POSNR', StringType(), True), StructField('PARVW', StringType(), True), StructField('PERNR', StringType(), True), StructField('LAND1', StringType(), True), StructField('ADRDA', StringType(), True), StructField('ASSIGNED_BP', StringType(), True), StructField('DUMMY_SDDOCPARTNER_INCL_EEW_PS', StringType(), True), StructField('SDM_VERSION', StringType(), True)]),

    "VBRK": StructType([StructField('MANDT', StringType(), True), StructField('VBELN', StringType(), True), 
                        StructField('FKART', StringType(), True), StructField('FKTYP', StringType(), True), StructField('VBTYP', StringType(), True), StructField('WAERK', StringType(), True), StructField('VKORG', StringType(), True), StructField('VTWEG', StringType(), True), StructField('KALSM', StringType(), True), StructField('KNUMV', StringType(), True), StructField('VSBED', StringType(), True), StructField('FKDAT', StringType(), True), StructField('BELNR', StringType(), True), StructField('GJAHR', StringType(), True), StructField('POPER', StringType(), True), StructField('KONDA', StringType(), True), StructField('BZIRK', StringType(), True), StructField('INCO1', StringType(), True), StructField('INCO2', StringType(), True), StructField('RFBSK', StringType(), True), StructField('KURRF', StringType(), True), StructField('CPKUR', StringType(), True), StructField('ZTERM', StringType(), True), StructField('ZLSCH', StringType(), True), StructField('KTGRD', StringType(), True), StructField('LAND1', StringType(), True), StructField('REGIO', StringType(), True), StructField('BUKRS', StringType(), True), StructField('TAXK1', StringType(), True), StructField('NETWR', StringType(), True), StructField('ZUKRI', StringType(), True), StructField('ERNAM', StringType(), True), StructField('ERZET', StringType(), True), StructField('ERDAT', StringType(), True), StructField('KUNRG', StringType(), True), StructField('KUNAG', StringType(), True), StructField('MABER', StringType(), True), StructField('STWAE', StringType(), True), StructField('FKART_RL', StringType(), True), StructField('KURST', StringType(), True), StructField('SPART', StringType(), True), StructField('KKBER', StringType(), True), StructField('KNKLI', StringType(), True), StructField('CMWAE', StringType(), True), StructField('CMKUF', StringType(), True), StructField('HITYP_PR', StringType(), True), StructField('BSTNK_VF', StringType(), True), StructField('KAPPL', StringType(), True), StructField('LANDTX', StringType(), True), StructField('STCEG_H', StringType(), True), StructField('STCEG_L', StringType(), True), StructField('XBLNR', StringType(), True), StructField('ZUONR', StringType(), True), StructField('MWSBK', StringType(), True), StructField('LOGSYS', StringType(), True), StructField('KURRF_DAT', StringType(), True), StructField('KIDNO', StringType(), True), StructField('CHANGED_ON', StringType(), True)]),

    "VBRP": StructType([StructField('ALAND', StringType(), True), StructField('ARKTX', StringType(), True), 
                        StructField('AUBEL', StringType(), True), StructField('AUPOS', StringType(), True), 
                        StructField('BONBA', StringType(), True), StructField('BRGEW', StringType(), True), 
                        StructField('CHARG', StringType(), True), StructField('EAN11', StringType(), True), 
                        StructField('ERDAT', StringType(), True), StructField('ERNAM', StringType(), True), 
                        StructField('ERZET', StringType(), True), StructField('FBUDA', StringType(), True), 
                        StructField('FKIMG', StringType(), True), StructField('FKLMG', StringType(), True), 
                        StructField('GEWEI', StringType(), True), StructField('KONDM', StringType(), True), 
                        StructField('KTGRM', StringType(), True), StructField('KURSK', StringType(), True), 
                        StructField('KVGR1', StringType(), True), StructField('KVGR2', StringType(), True), 
                        StructField('KVGR3', StringType(), True), StructField('KVGR4', StringType(), True), 
                        StructField('KVGR5', StringType(), True), StructField('KZWI1', StringType(), True), 
                        StructField('KZWI2', StringType(), True), StructField('KZWI3', StringType(), True), 
                        StructField('KZWI4', StringType(), True), StructField('KZWI5', StringType(), True), 
                        StructField('KZWI6', StringType(), True), StructField('LGORT', StringType(), True), 
                        StructField('LMENG', StringType(), True), StructField('MANDT', StringType(), True), 
                        StructField('MATKL', StringType(), True), StructField('MATNR', StringType(), True), 
                        StructField('MATWA', StringType(), True), StructField('MEINS', StringType(), True), 
                        StructField('MVGR1', StringType(), True), StructField('MVGR2', StringType(), True), 
                        StructField('MVGR3', StringType(), True), StructField('MVGR4', StringType(), True), 
                        StructField('MVGR5', StringType(), True), StructField('NETWR', StringType(), True), 
                        StructField('NTGEW', StringType(), True), StructField('PMATN', StringType(), True), 
                        StructField('POSNR', StringType(), True), StructField('POSNV', StringType(), True), 
                        StructField('POSPA', StringType(), True), StructField('PRCTR', StringType(), True), 
                        StructField('PRODH', StringType(), True), StructField('PRSDT', StringType(), True), 
                        StructField('PRSFD', StringType(), True), StructField('PSTYV', StringType(), True), 
                        StructField('SKFBP', StringType(), True), StructField('SKTOF', StringType(), True), 
                        StructField('SMENG', StringType(), True), StructField('SPARA', StringType(), True), 
                        StructField('SPART', StringType(), True), StructField('STAFO', StringType(), True), 
                        StructField('STCUR', StringType(), True), StructField('TAXM1', StringType(), True), 
                        StructField('UMVKN', StringType(), True), StructField('UMVKZ', StringType(), True), 
                        StructField('VBELN', StringType(), True), StructField('VBELV', StringType(), True), 
                        StructField('VGBEL', StringType(), True), StructField('VGPOS', StringType(), True), 
                        StructField('VGTYP', StringType(), True), StructField('VKBUR', StringType(), True), 
                        StructField('VKGRP', StringType(), True), StructField('VOLEH', StringType(), True), 
                        StructField('VOLUM', StringType(), True), StructField('VRKME', StringType(), True), 
                        StructField('VSTEL', StringType(), True), StructField('WAVWR', StringType(), True), 
                        StructField('WERKS', StringType(), True), StructField('WKREG', StringType(), True)]),
    
    "VBEP" :  StructType([StructField('MANDT', StringType(), True), StructField('VBELN', StringType(), True), 
                          StructField('POSNR', StringType(), True), StructField('ETENR', StringType(), True), 
                          StructField('ETTYP', StringType(), True), StructField('LFREL', StringType(), True), 
                          StructField('EDATU', StringType(), True), StructField('EZEIT', StringType(), True), 
                          StructField('WMENG', StringType(), True), StructField('BMENG', StringType(), True), 
                          StructField('VRKME', StringType(), True), StructField('LMENG', StringType(), True), 
                          StructField('MEINS', StringType(), True), StructField('BDART', StringType(), True), 
                          StructField('PLART', StringType(), True), StructField('IDNNR', StringType(), True), 
                          StructField('PRGRS', StringType(), True), StructField('TDDAT', StringType(), True), 
                          StructField('MBDAT', StringType(), True), StructField('LDDAT', StringType(), True), 
                          StructField('WADAT', StringType(), True), StructField('CMENG', StringType(), True), 
                          StructField('ROMS1', StringType(), True), StructField('UMVKZ', StringType(), True), 
                          StructField('UMVKN', StringType(), True), StructField('VERFP', StringType(), True), 
                          StructField('BWART', StringType(), True), StructField('ABGES', StringType(), True), 
                          StructField('MBUHR', StringType(), True), StructField('TDUHR', StringType(), True), 
                          StructField('LDUHR', StringType(), True), StructField('WAUHR', StringType(), True), 
                          StructField('HANDOVERTIME', StringType(), True), StructField('DLVQTY_BU', StringType(), True), StructField('DLVQTY_SU', StringType(), True), StructField('OCDQTY_BU', StringType(), True), StructField('OCDQTY_SU', StringType(), True), StructField('ORDQTY_BU', StringType(), True), StructField('ORDQTY_SU', StringType(), True), StructField('CREA_DLVDATE', StringType(), True), StructField('REQ_DLVDATE', StringType(), True), StructField('BEDAR', StringType(), True), StructField('WAERK', StringType(), True), StructField('ODN_AMOUNT', StringType(), True), StructField('HANDLE', StringType(), True), StructField('LCCST', StringType(), True), 
                          StructField('RRQQTY_BU', StringType(), True), StructField('CRQQTY_BU', StringType(), True), StructField('DUMMY_SLSDOCSCHEDL_INCL_EEW_PS', StringType(), True), 
                          StructField('FSH_RALLOC_QTY', StringType(), True)])
    }
    if table == False:
        return schema_dict.keys()
    else:
        return schema_dict[table]

# COMMAND ----------

def save_excel_files(data_dict):
    for nombre_del_archivo, df in data_dict.items():
        # Ruta del archivo Excel
        excel_path = f"/tmp/{nombre_del_archivo}.xlsx"
        
        # Convertir PySpark DataFrame a Pandas DataFrame
        df_pandas = df.toPandas()
        
        # Crear un writer de Pandas usando XlsxWriter
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        
        # Convertir el DataFrame a Excel y guardar localmente
        df_pandas.to_excel(writer, index=False)
        writer.save()
        
        # Mover el archivo al punto de montaje en DBFS (si estás en Databricks)
        dbutils.fs.cp(f"file:{excel_path}",
                      f"/mnt/adls-dac-data-bi-pr-scus/powerappsentregables/{nombre_del_archivo}.xlsx")
        
        print(f"Archivo guardado: {nombre_del_archivo}.xlsx")

# COMMAND ----------

def process_user(mail, user_type, demanda_base_mi):
    id = mail.replace("@kuoafmkt.com", "").replace(".", "_")
    filename = f"baseescalera_{id}_{current_year}_{current_month}.xlsx"
    temp_path = f"{temp_local_path}/{filename}"
    final_path = f"{base_path}/{filename}"
    
    if user_type == "esp":
        paths.append(final_path)
        df = demanda_base_mi.filter(col("ESPECIALISTA") == mail).orderBy(col("INDICADOR_ABC").asc())
        df = df.toPandas()

        #Pestaña resumen

        df_resumen = (demanda_base_mi.filter(col("ESPECIALISTA") == mail)
                            .select("SISTEMA", "LINEA_AGRUPADA")
                            .dropDuplicates()
                            .orderBy(col("SISTEMA").asc()))            
            
        cols = demanda_base_mi.columns[29:]
        for c in cols:
            df_resumen = df_resumen.withColumn(c, lit(""))

        df_resumen = df_resumen.toPandas()
        
    else:
        paths.append(final_path)
        df = demanda_base_mi.orderBy(col("INDICADOR_ABC").asc())
        df = df.toPandas()

        #Pestaña resumen

        df_resumen = (demanda_base_mi
                            .select("SISTEMA", "LINEA_AGRUPADA")
                            .dropDuplicates()
                            .orderBy(col("SISTEMA").asc()))            
            
        cols = demanda_base_mi.columns[29:]
        for c in cols:
            df_resumen = df_resumen.withColumn(c, lit(""))

        df_resumen = df_resumen.toPandas()
    
    # Inicialización de variables para los rangos dinámicos
    inicio_prono_unidades = None
    inicio_prono_importe = None
    inicio_prono_variacion = None

    current_month_str = f"{current_month:02d}"
    
    # Búsqueda de las columnas de inicio de prono unidades e importes
    for index, column in enumerate(df.columns):
        if column == f"{current_year}-01 U PRONO":
            inicio_prono_unidades = index + 1
        if column == f"{current_year}-01 $ PRONO":
            inicio_prono_importe = index + 1
        if column == f"{current_year}-{current_month_str} U PRONO": #2024-05
            inicio_prono_variacion = index + 1
    
    # Verificar si las columnas fueron encontradas
    if inicio_prono_unidades is None:
        raise ValueError(f"Columna {current_year}-01 U PRONO no encontrada en el DataFrame")
    if inicio_prono_importe is None:
        raise ValueError(f"Columna {current_year}-01 $ PRONO no encontrada en el DataFrame")
    if inicio_prono_variacion is None:
        raise ValueError(f"Columna {current_year}-{current_month} U PRONO no encontrada en el DataFrame")
    
    # Cálculo de los rangos dinámicos
    inicio_llenado = inicio_prono_unidades + current_month + 2
    fin_llenado =  inicio_llenado + 18
    fin_prono_unidades = inicio_prono_unidades + 32
    fin_prono_importe = inicio_prono_importe + 32
    fin_prono_variacion12 = inicio_prono_variacion + 11
    fin_prono_variacion6 = inicio_prono_variacion + 5
    fin_prono_variacion3 = inicio_prono_variacion + 2

    workbook = pd.ExcelWriter(temp_path, engine='openpyxl')
    df.to_excel(workbook, sheet_name="BASE", startrow=0, index=False, header=True)
    df_resumen.to_excel(workbook, sheet_name="RESUMEN", startrow=0, index=False, header=True)
    workbook.save()
    
    wb = load_workbook(temp_path)
    ws = wb['BASE']
    wsr = wb['RESUMEN']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for c in range(30, 185):
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            cell.number_format = '0.00'  # Formato de número con dos decimales


    for column in range(inicio_llenado, fin_llenado):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).fill = yellow_fill

    for row in range(2, len(df) + 2):  # filas
        promocion = ws.cell(row=row, column=19)
        promocion_str = ws.cell(row=row, column=19).coordinate
        venta = ws.cell(row=row, column=20)
        venta_str = ws.cell(row=row, column=20).coordinate
        precio_de_lista = ws.cell(row=row, column=15)
        precio_de_lista_str = ws.cell(row=row, column=15).coordinate
        
        # Rango dinámico de columnas prono unidades e importe 
        for c, c1 in zip(range(inicio_prono_unidades, fin_prono_unidades),
                         range(inicio_prono_importe, fin_prono_importe)):
            unidades = ws.cell(row=row, column=c).coordinate

            ws.cell(row=row, column=c1).value = (
                f'=ROUND(IF({promocion_str}>0,({unidades}*{promocion_str})/1000,({unidades}*{venta_str})/1000),2)'
            )
            
        ws.cell(row=row, column=24).value = f'=ROUND(AVERAGE(BB{row}:BM{row}),2)' #PROMEDIO 12M
        ws.cell(row=row, column=25).value = f'=ROUND(AVERAGE(BH{row}:BM{row}),2)' #PROMEDIO 6M
        ws.cell(row=row, column=26).value = f'=ROUND(AVERAGE(BK{row}:BM{row}),2)' #PROMEDIO 3M

        celda_inicio = ws.cell(row=row, column=inicio_prono_variacion).coordinate #Celda inicio rango var 12m
        celda_fin = ws.cell(row=row, column=fin_prono_variacion12).coordinate #Celda fin rango var 12m
        ws.cell(row=row, column=27).value = f'=ROUND(X{row}/(AVERAGE({celda_inicio}:{celda_fin}))*100,2)'
        
        celda_inicio = ws.cell(row=row, column=inicio_prono_variacion).coordinate #Celda inicio rango var 6m
        celda_fin = ws.cell(row=row, column=fin_prono_variacion6).coordinate #Celda fin rango var 6m
        ws.cell(row=row, column=28).value = f'=ROUND(Y{row}/(AVERAGE({celda_inicio}:{celda_fin}))*100,2)'

        celda_inicio = ws.cell(row=row, column=inicio_prono_variacion).coordinate #Celda inicio rango var 3m
        celda_fin = ws.cell(row=row, column=fin_prono_variacion3).coordinate #Celda fin rango var 3m
        ws.cell(row=row, column=29).value = f'=ROUND(Z{row}/(AVERAGE({celda_inicio}:{celda_fin}))*100,2)'
        
        ws.cell(row=row, column=67).value = f'=ROUND(SUM(BP{row}:CA{row}),2)' #Acumulado prono 24
    
        f = len(df) + 1
        crit_1_in = ws.cell(row=2, column=3).coordinate
        crit_2_in = ws.cell(row=2, column=6).coordinate
        crit_1_fin = ws.cell(row=f, column=3).coordinate
        crit_2_fin = ws.cell(row=f, column=6).coordinate
        for row in range(2, len(df_resumen) + 2):
            for c, c1 in zip(range(30, 192), range(3, 165)): #RANGO PESTAÑA BASE, RANGO PESTAÑA RESUMEN
                # La coordenada de las celdas A y B debe actualizarse para cada fila
                c_inicio = ws.cell(row=2, column=c).coordinate
                c_fin = ws.cell(row=f, column=c).coordinate
                crit_1_resumen = wsr.cell(row=row, column=1).coordinate
                crit_2_resumen = wsr.cell(row=row, column=2).coordinate
                wsr.cell(row=row, column=c1).value = (
                    f"=SUMIFS(BASE!{c_inicio}:{c_fin},BASE!{crit_1_in}:{crit_1_fin},{crit_1_resumen},BASE!{crit_2_in}:{crit_2_fin},{crit_2_resumen})"
                )


    wb.save(temp_path)
    wb.close()
    dbutils.fs.cp(f'file:{temp_path}', "/mnt/" + final_path)
    print(f"Archivo guardado en: {final_path}")


# COMMAND ----------

def process_user_me(mail, user_type, df_input):
    id = mail.replace("@kuoafmkt.com", "").replace(".", "_")
    filename = f"baseescalera_{id}_{current_year}_{current_month}.xlsx"
    temp_path = f"{temp_local_path}/{filename}"
    final_path = f"{base_path}/{filename}"
    
    if user_type == "ases":
        paths.append(final_path)
        df = df_input.orderBy(col("RAZON_SOCIAL").asc())
        df = df.toPandas()

        #Pestaña resumen

        df_resumen = (df_input
                            .select("ID_CANAL_DISTRIBUCION","SISTEMA", "LINEA_AGRUPADA")
                            .dropDuplicates()
                            .orderBy(col("SISTEMA").asc()))
        
        cols = df_input.columns[27:]
        for c in cols:
            df_resumen = df_resumen.withColumn(c, lit(""))

        df_resumen = df_resumen.toPandas()
           
    else:
        paths.append(final_path)
        df = df_input.orderBy(col("RAZON_SOCIAL").asc())
        df = df.toPandas()

    #Pestaña resumen

        df_resumen = (df_input
                            .select("ID_CANAL_DISTRIBUCION","SISTEMA", "LINEA_AGRUPADA")
                            .dropDuplicates()
                            .orderBy(col("SISTEMA").asc()))            
            
        cols = df_input.columns[27:]
        for c in cols:
            df_resumen = (df_resumen.withColumn(c, lit(""))
                          .orderBy("SISTEMA","LINEA_AGRUPADA","ID_CANAL_DISTRIBUCION")
                          )

        df_resumen = df_resumen.toPandas()
    
    # Inicialización de variables para los rangos dinámicos
    inicio_prono_unidades = None
    inicio_prono_importe = None
    inicio_prono_variacion = None

    current_month_str = f"{current_month:02d}"
    
    # Búsqueda de las columnas de inicio de prono unidades e importes
    for index, column in enumerate(df.columns):
        if column == f"{current_year}-01 U PRONO":
            inicio_prono_unidades = index + 1
        if column == f"{current_year}-01 $ PRONO":
            inicio_prono_importe = index + 1
        # if column == f"{current_year}-{current_month_str} U PRONO": #2024-05
        #     inicio_prono_variacion = index + 1
    
    # Verificar si las columnas fueron encontradas
    if inicio_prono_unidades is None:
        raise ValueError(f"Columna {current_year}-01 U PRONO no encontrada en el DataFrame")
    if inicio_prono_importe is None:
        raise ValueError(f"Columna {current_year}-01 $ PRONO no encontrada en el DataFrame")
    # if inicio_prono_variacion is None:
    #     raise ValueError(f"Columna {current_year}-{current_month} U PRONO no encontrada en el DataFrame")
    
    # Cálculo de los rangos dinámicos
    inicio_llenado = inicio_prono_unidades + current_month + 2
    fin_llenado =  inicio_llenado + 18
    fin_prono_unidades = inicio_prono_unidades + 32
    fin_prono_importe = inicio_prono_importe + 32
    # fin_prono_variacion12 = inicio_prono_variacion + 11
    # fin_prono_variacion6 = inicio_prono_variacion + 5
    # fin_prono_variacion3 = inicio_prono_variacion + 2

    workbook = pd.ExcelWriter(temp_path, engine='openpyxl')
    df.to_excel(workbook, sheet_name="BASE", startrow=0, index=False, header=True)
    df_resumen.to_excel(workbook, sheet_name="RESUMEN", startrow=0, index=False, header=True)
    workbook.save()
    
    wb = load_workbook(temp_path)
    ws = wb['BASE']
    wsr = wb['RESUMEN']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for c in range(30, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            cell.number_format = '0.00'  # Formato de número con dos decimales

    for column in range(inicio_llenado, fin_llenado):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).fill = yellow_fill

    for row in range(2, len(df) + 2):  # filas
        precio_de_lista = ws.cell(row=row, column=22)
        precio_de_lista_str = ws.cell(row=row, column=22).coordinate
        
        # Rango dinámico de columnas prono unidades e importe 
        for c, c1 in zip(range(inicio_prono_unidades, fin_prono_unidades),
                         range(inicio_prono_importe, fin_prono_importe)):
            unidades = ws.cell(row=row, column=c).coordinate

            ws.cell(row=row, column=c1).value = (
                f'=ROUND(({unidades}*{precio_de_lista_str})/1000,2)'
            )
            
        # ws.cell(row=row, column=23).value = f'=ROUND(AVERAGE(BB{row}:BM{row}),2)' #PROMEDIO 12M
        # ws.cell(row=row, column=24).value = f'=ROUND(AVERAGE(BH{row}:BM{row}),2)' #PROMEDIO 6M
        # ws.cell(row=row, column=25).value = f'=ROUND(AVERAGE(BK{row}:BM{row}),2)' #PROMEDIO 3M

        # celda_inicio = ws.cell(row=row, column=inicio_prono_variacion).coordinate #Celda inicio rango var 12m
        # celda_fin = ws.cell(row=row, column=fin_prono_variacion12).coordinate #Celda fin rango var 12m
        # ws.cell(row=row, column=26).value = f'=ROUND(W{row}/(AVERAGE({celda_inicio}:{celda_fin}))*100,2)'
        
        # celda_inicio = ws.cell(row=row, column=inicio_prono_variacion).coordinate #Celda inicio rango var 6m
        # celda_fin = ws.cell(row=row, column=fin_prono_variacion6).coordinate #Celda fin rango var 6m
        # ws.cell(row=row, column=27).value = f'=ROUND(X{row}/(AVERAGE({celda_inicio}:{celda_fin}))*100,2)'

        # celda_inicio = ws.cell(row=row, column=inicio_prono_variacion).coordinate #Celda inicio rango var 3m
        # celda_fin = ws.cell(row=row, column=fin_prono_variacion3).coordinate #Celda fin rango var 3m
        # ws.cell(row=row, column=28).value = f'=ROUND(Y{row}/(AVERAGE({celda_inicio}:{celda_fin}))*100,2)'
        
        ws.cell(row=row, column=64).value = f'=ROUND(SUM(BM{row}:BX{row}),2)' #Acumulado prono 24

        f = len(df) + 1
        crit_1_in = ws.cell(row=2, column=10).coordinate
        crit_2_in = ws.cell(row=2, column=13).coordinate
        crit_3_in = ws.cell(row=2, column=17).coordinate
        crit_1_fin = ws.cell(row=f, column=10).coordinate
        crit_2_fin = ws.cell(row=f, column=13).coordinate
        crit_3_fin = ws.cell(row=f,column=17).coordinate

        for row in range(2, len(df_resumen) + 2):
            for c, c1 in zip(range(28, 164), range(4, 139)):#RANGO PESTAÑA BASE, RANGO PESTAÑA RESUMEN
                # La coordenada de las celdas A y B debe actualizarse para cada fila
                c_inicio = ws.cell(row=2, column=c).coordinate
                c_fin = ws.cell(row=f, column=c).coordinate
                crit_1_resumen = wsr.cell(row=row, column=2).coordinate
                crit_2_resumen = wsr.cell(row=row, column=3).coordinate
                crit_3_resumen = wsr.cell(row=row, column=1).coordinate
                wsr.cell(row=row, column=c1).value = (
                    f"=SUMIFS(BASE!{c_inicio}:{c_fin},BASE!{crit_1_in}:{crit_1_fin},{crit_1_resumen},BASE!{crit_2_in}:{crit_2_fin},{crit_2_resumen},BASE!{crit_3_in}:{crit_3_fin},{crit_3_resumen})"
                )

    wb.save(temp_path)
    wb.close()
    dbutils.fs.cp(f'file:{temp_path}', "/mnt/" + final_path)
    print(f"Archivo guardado en: {final_path}")


# COMMAND ----------

def union_multiple_dataframes(*dfs: DataFrame) -> DataFrame:
    if not dfs:
        raise ValueError("Se debe proporcionar al menos un DataFrame")

    # Obtener el conjunto de todas las columnas presentes en cualquier DataFrame
    all_columns = set()
    for df in dfs:
        all_columns.update(df.columns)

    # Añadir las columnas faltantes a cada DataFrame con valores nulos
    filled_dfs = []
    for df in dfs:
        missing_columns = all_columns - set(df.columns)
        for col in missing_columns:
            df = df.withColumn(col, lit(None))
        filled_dfs.append(df.select(sorted(all_columns)))

    # Realizar la unión de todos los DataFrames
    result_df = filled_dfs[0]
    for df in filled_dfs[1:]:
        result_df = result_df.unionByName(df,allowMissingColumns = True)

    return result_df

# COMMAND ----------

def process_escalera(esc_mi, esc_mi_sis_lin,total_fritec,total_frenos,total_fritec_frenos,
                     total_motor, total_tren_motriz, esc_me, esc_mo, total_mi_df, total_memo_df, total_dacomsa_df):
    filename = f"escalera_{current_year}_{current_month}.xlsx"
    temp_path = f"{temp_local_path}/{filename}"
    final_path = f"{base_path}/{filename}"
    row_count_mi = 0
    row_count_memo = 0

    #Bloque de dataframes de MI
    with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
        for sis_lin  in esc_mi_sis_lin:
            print("-------------------------------------------")
            df = (esc_mi.filter(col("SISTEMA-LINEA_AGRUPADA") == sis_lin)
                  .withColumn("TOTAL ANUAL",lit(""))
                  .withColumn("PROMEDIO ANUAL",lit(""))
                  )
            print(f"Escribiendo: {df.select('SISTEMA-LINEA_AGRUPADA').distinct().rdd.flatMap(lambda x: x).collect()}")
            df = df.drop("SISTEMA-LINEA_AGRUPADA")
            row_ppto = df.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
            row_revisado = df.count() + 1 + row_count_mi 
            df = df.toPandas()
            df.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi, index=False, header=True)
            ws = writer.sheets["ESC-MI"]
            print("Escribiendo fórmulas de TOTAL y PROMEDIO ANUAL")
            for r in range(row_count_mi + 2, row_count_mi + len(df) + 2):
                ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
                ws.cell(row=r, column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'

            formula_ppto_prom_tri = row_count_mi + len(df) + 2
            formula_revisado_prom_tri = row_count_mi + len(df) + 3
            formula_alc_vs_ppto = row_count_mi + len(df) + 4

            print("Escribiendo fórmulas de promedios trimestrales de ppto y revisado ...")
            for c in (6,9,12,15):
                c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
                c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
                c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
                c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
                ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
                ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
                ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
                revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
                ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'
                

            ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
            ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
            ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"

            print(f"Celda ppto:{c_inicio_ppto}")
            print(f"Celda revisado:{c_inicio_revisado}")

            row_count_mi += len(df) + 6

            print(f"Cuenta de fila para pestaña MI: {row_count_mi}")

        rango_sumif_max = row_count_mi + 2
        print(F"Fila de máxima para el rango de SUMIF de totales por sistema: {rango_sumif_max}")

        #Dataframe total fritec
        print("-------------------------------------------")
        print("Escribiendo: TOTAL FRITEC")
        row_ppto = total_fritec.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
        row_revisado = total_fritec.count() + 1 + row_count_mi
        df_total_fritec = (total_fritec
                            .withColumn("TOTAL ANUAL",lit(""))
                           .withColumn("PROMEDIO ANUAL",lit(""))                          
                           .toPandas()
                           )
        df_total_fritec.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi,index=False, header=True)
        ws = writer.sheets["ESC-MI"]
        for r in range(row_count_mi + 2, row_count_mi + len(df_total_fritec) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df_total_fritec) + 3
        formula_revisado_prom_tri = row_count_mi + len(df_total_fritec) + 4
        formula_alc_vs_ppto = row_count_mi + len(df_total_fritec) + 5

        print("Escribeindo fórmulas de promedios trimestrales de ppto y revisado ...")
        for c in (6,9,12,15):
            c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
            ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
            ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
            revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
            ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"

        print("Escribeindo fórmulas de de SUMIF de totales FRITEC...")
        for c in range(4,36):
            celda_inicio = ws.cell(row = 2, column=c).coordinate
            celda_fin = ws.cell(row = rango_sumif_max, column=c).coordinate
            print(f"Rango de SUMIF - FRITEC:  {celda_inicio}:{celda_fin}")
            ws.cell(row=row_revisado, column=c).value = f'=SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BIOCERAMIC FT") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BALATA FRENO TAMBOR") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BALATA FRENO DISCO")'


        print("TOTAL-FRITEC terminado...")
        print(f"Celda ppto FRITEC:{c_inicio_ppto}")
        print(f"Celda revisado FRITEC:{c_inicio_revisado}")

        row_count_mi += len(df_total_fritec) + 7

        #Dataframe total frenos
        print("-------------------------------------------")
        print("Escribiendo: TOTAL Frenos")
        row_ppto = total_frenos.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
        row_revisado = total_frenos.count() + 1 + row_count_mi
        df_total_frenos = (total_frenos
                            .withColumn("TOTAL ANUAL",lit(""))
                            .withColumn("PROMEDIO ANUAL",lit(""))
                            .toPandas()
                           )
        df_total_frenos.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi,index=False, header=True)
        ws = writer.sheets["ESC-MI"]
        for r in range(row_count_mi + 2, row_count_mi + len(df_total_frenos) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df_total_frenos) + 3
        formula_revisado_prom_tri = row_count_mi + len(df_total_frenos) + 4
        formula_alc_vs_ppto = row_count_mi + len(df_total_frenos) + 5

        print("Escribeindo fórmulas de promedios trimestrales de ppto y revisado ...")
        for c in (6,9,12,15):
            c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
            ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
            ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
            revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
            ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"


        print("Escribeindo fórmulas de de SUMIF de totales Frenos...")
        for c in range(4,36):
            celda_inicio = ws.cell(row = 2, column=c).coordinate
            celda_fin = ws.cell(row = rango_sumif_max, column=c).coordinate
            print(f"Rango de SUMIF - FRENOS:  {celda_inicio}:{celda_fin}")
            ws.cell(row=row_revisado, column=c).value = f'=SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"FRITEC ROTORES") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BALATAS BIOCERAMIC") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"RACE MAZAS RUEDA") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"FRITEC BLOCKS")'

        print("TOTAL-FRENOS terminado ..")
        print(f"Celda ppto:{c_inicio_ppto}")
        print(f"Celda revisado:{c_inicio_revisado}")

        row_count_mi += len(df_total_frenos) + 7

        inicio_sumif_totales = row_count_mi

        print(f"Fila de incio de rango para el SUMIF para TOTAL-MI: {inicio_sumif_totales}")

        #Dataframe total fritec + frenos
        print("-------------------------------------------")
        print("Escribiendo: TOTAL FRITEC + FRENOS")
        row_ppto = total_fritec_frenos.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
        row_revisado = total_fritec_frenos.count() + 1 + row_count_mi
        df_total_fritec_frenos= (total_fritec_frenos
                                 .withColumn("TOTAL ANUAL",lit(""))
                                 .withColumn("PROMEDIO ANUAL",lit(""))
                                 .toPandas()
                                 )
        df_total_fritec_frenos.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi,index=False, header=True)
        ws = writer.sheets["ESC-MI"]
        for r in range(row_count_mi + 2, row_count_mi + len(df_total_fritec_frenos) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df_total_fritec_frenos) + 3
        formula_revisado_prom_tri = row_count_mi + len(df_total_fritec_frenos) + 4
        formula_alc_vs_ppto = row_count_mi + len(df_total_fritec_frenos) + 5

        print("Escribeindo fórmulas promedio trimestrales de ppto y revisado ...")

        for c in (6,9,12,15):
            c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
            ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
            ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
            revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
            ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"

        print("Escribiendo fórmulas de SUMIF de totales FRENOS + FRITEC...")
        for c in range(4,36):
            celda_inicio = ws.cell(row = 2, column=c).coordinate
            celda_fin = ws.cell(row = rango_sumif_max, column=c).coordinate
            print(f"Rango de SUMIF - FRENOS:  {celda_inicio}:{celda_fin}")
            ws.cell(row=row_revisado, column=c).value = f'=SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"FRITEC ROTORES") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BALATAS BIOCERAMIC") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"RACE MAZAS RUEDA") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"FRITEC BLOCKS") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BIOCERAMIC FT") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BALATA FRENO TAMBOR") + SUMIFS({celda_inicio}:{celda_fin},$A$2:$A${rango_sumif_max},"REVISADO",$B$2:$B${rango_sumif_max},"BALATA FRENO DISCO")'

        print("TOTAL-FRITEC-FRENOS terminado ...")
        print(f"Celda ppto FRITEC-FRENOS: {c_inicio_ppto}")
        print(f"Celda revisado FRITEC-FRENOS: {c_inicio_revisado}")

        row_count_mi += len(df_total_fritec_frenos) + 7


        #Dataframe total motor
        print("-------------------------------------------")
        print("Escribiendo: TOTAL MOTOR")        
        row_ppto = total_motor.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
        row_revisado = total_motor.count() + 1 + row_count_mi
        df_total_motor = (total_motor
                          .withColumn("TOTAL ANUAL",lit(""))
                          .withColumn("PROMEDIO ANUAL",lit(""))
                          .toPandas()
                          )
        df_total_motor.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi,index=False, header=True)
        for r in range(row_count_mi + 2, row_count_mi + len(df_total_motor) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df_total_motor) + 3
        formula_revisado_prom_tri = row_count_mi + len(df_total_motor) + 4
        formula_alc_vs_ppto = row_count_mi + len(df_total_motor) + 5


        print("Escribeindo fórmulas promedio trimestrales de ppto y revisado ...")
        for c in (6,9,12,15):
            c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
            ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
            ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
            revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
            ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"

        print("Escribiendo fórmulas de SUMIF de totales MOTOR...")
        for c in range(4,36):
            celda_inicio = ws.cell(row = 2, column=c).coordinate
            celda_fin = ws.cell(row = rango_sumif_max, column=c).coordinate
            print(f"Rango de SUMIF - MOTOR:  {celda_inicio}:{celda_fin}")
            ws.cell(row=row_revisado, column=c).value = f'=SUMIFS({celda_inicio}:{celda_fin},$C$2:$C${rango_sumif_max},"MOTOR",$A2:$A${rango_sumif_max},"REVISADO")'

        print("TOTAL-MOTOR terminado ...")
        print(f"Celda ppto MOTOR:{c_inicio_ppto}")
        print(f"Celda revisado MOTOR:{c_inicio_revisado}") 

        row_count_mi += len(df_total_motor) + 7

        #Dataframe total tren motriz

        print("-------------------------------------------")
        print("Escribiendo: TOTAL TREN MOTRIZ")    
        row_ppto = total_tren_motriz.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
        row_revisado = total_tren_motriz.count() + 1 + row_count_mi
        df_total_tren_motriz = (total_tren_motriz
                                .withColumn("TOTAL ANUAL",lit(""))
                                .withColumn("PROMEDIO ANUAL",lit(""))
                                .toPandas()
                                )
        df_total_tren_motriz.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi,index=False, header=True)
        ws = writer.sheets["ESC-MI"]
        for r in range(row_count_mi + 2, row_count_mi + len(df_total_tren_motriz) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df_total_tren_motriz) + 3
        formula_revisado_prom_tri = row_count_mi + len(df_total_tren_motriz) + 4
        formula_alc_vs_ppto = row_count_mi + len(df_total_tren_motriz) + 5

        print("Escribeindo fórmulas promedio trimestrales de ppto y revisado ...")

        for c in (6,9,12,15):
            c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
            ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
            ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
            revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
            ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"


        print("Escribiendo fórmulas de SUMIF de totales TREN MOTRIZ...")
        for c in range(4,36):
            celda_inicio = ws.cell(row = 2, column=c).coordinate
            celda_fin = ws.cell(row = rango_sumif_max, column=c).coordinate
            print(f"Rango de SUMIF - TREN MOTRIZ:  {celda_inicio}:{celda_fin}")
            ws.cell(row=row_revisado, column=c).value = f'=SUMIFS({celda_inicio}:{celda_fin},$C$2:$C${rango_sumif_max},"TREN MOTRIZ",$A$2:$A${rango_sumif_max},"REVISADO")'

        print("TOTAL-TREN MOTRIZ terminado ..")
        print(f"Celda ppto TREN MOTRIZ:{c_inicio_ppto}")
        print(f"Celda revisado TREN MOTRIZ:{c_inicio_revisado}")

        row_count_mi += len(df_total_tren_motriz) + 6

        fin_sumifs_totales = row_count_mi - 2

        print(f"Fin del rango del SUMIF de los totales: {fin_sumifs_totales}")

        #Dataframe total MI

        print("-------------------------------------------")
        print("Escribiendo: TOTAL MI")    
        row_ppto = total_mi_df.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_mi
        row_revisado = total_mi_df.count() + 1 + row_count_mi
        df_total_mi = (total_mi_df
                       .withColumn("TOTAL ANUAL",lit(""))
                        .withColumn("PROMEDIO ANUAL",lit(""))
                        .toPandas()
                        )
        df_total_mi.to_excel(writer, sheet_name="ESC-MI", startrow=row_count_mi,index=False, header=True)
        ws = writer.sheets["ESC-MI"]
        for r in range(row_count_mi + 2, row_count_mi + len(df_total_mi) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df_total_mi) + 3
        formula_revisado_prom_tri = row_count_mi + len(df_total_mi) + 4
        formula_alc_vs_ppto = row_count_mi + len(df_total_mi) + 5

        for c in (6,9,12,15):
            c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = ws.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
            ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
            ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
            revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
            ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        ws.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        ws.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        ws.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"

        for c in range(4,36):
            celda_inicio = ws.cell(row = inicio_sumif_totales, column=c).coordinate
            celda_fin = ws.cell(row = fin_sumifs_totales, column=c).coordinate
            print(f"Rango de SUMIF - TOTAL MI:  {celda_inicio}:{celda_fin}")
            ws.cell(row=row_revisado, column=c).value = f'=SUMIFS({celda_inicio}:{celda_fin},$B${inicio_sumif_totales}:$B${fin_sumifs_totales},"TOTAL",$A${inicio_sumif_totales}:$A${fin_sumifs_totales},"REVISADO")'

        print("TOTAL-MI terminado...")
        print(f"Celda ppto total mi:{c_inicio_ppto}")
        print(f"Celda revisado total mi:{c_inicio_revisado}")
        print("-------------------------------------------")
        print(f"Escribiendo la pestña de ESC-MEMO")
        print("-------------------------------------------")
        print("Escribiendo Total ME....")

        #Dataframe total ME
        row_ppto = esc_me.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_memo
        row_revisado = esc_me.count() + 1 + row_count_memo
        df_me = (esc_me
                 .withColumn("TOTAL ANUAL",lit(""))
                 .withColumn("PROMEDIO ANUAL",lit(""))
                 .toPandas()
                )
        df_me.to_excel(writer, sheet_name="ESC-MEMO", startrow=row_count_memo, index=False, header=True)
        wsme = writer.sheets["ESC-MEMO"]
        for r in range(row_count_memo + 2, row_count_memo + len(df_me) + 2):
            wsme.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            wsme.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        # formula_ppto_prom_tri = row_count_mi + len(df) + 2
        # formula_revisado_prom_tri = row_count_mi + len(df) + 3
        # formula_alc_vs_ppto = row_count_mi + len(df) + 4

        # for c in (6,9,12,15):
        #     c_inicio_ppto = wsme.cell(row = row_ppto, column= c - 2).coordinate
        #     c_fin_ppto = wsme.cell(row = row_ppto, column= c).coordinate
        #     c_inicio_revisado = wsme.cell(row = row_revisado, column= c - 3).coordinate
        #     c_fin_revisado = wsme.cell(row = row_revisado, column= c).coordinate
        #     wsme.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
        #     ppto = wsme.cell(row=formula_ppto_prom_tri, column=c).coordinate
        #     wsme.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
        #     revisado = wsme.cell(row=formula_revisado_prom_tri, column=c).coordinate
        #     wsme.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'



        row_count_memo += len(df_me) + 3

        print("---------------------------------")
        print(f"Cuenta de fila de pestaña MEMO: {row_count_memo}")
        print("Escribiendo Total MO....")

        #Dataframe total MO
        row_ppto = esc_mo.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_memo
        row_revisado = esc_mo.count() + 1 + row_count_memo
        df_mo = (esc_mo
                 .withColumn("TOTAL ANUAL",lit(""))
                 .withColumn("PROMEDIO ANUAL",lit(""))
                 .toPandas()
                )
        df_mo.to_excel(writer, sheet_name="ESC-MEMO", startrow=row_count_memo, index=False, header=True)
        wsme = writer.sheets["ESC-MEMO"]
        for r in range(row_count_memo + 2, row_count_memo + len(df_mo) + 2):
            wsme.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            wsme.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        # formula_ppto_prom_tri = row_count_mi + len(df) + 2
        # formula_revisado_prom_tri = row_count_mi + len(df) + 3
        # formula_alc_vs_ppto = row_count_mi + len(df) + 4

        # for c in (6,9,12,15):
        #     c_inicio_ppto = wsme.cell(row = row_ppto, column= c - 2).coordinate
        #     c_fin_ppto = wsme.cell(row = row_ppto, column= c).coordinate
        #     c_inicio_revisado = wsme.cell(row = row_revisado, column= c - 2).coordinate
        #     c_fin_revisado = wsme.cell(row = row_revisado, column= c).coordinate
        #     wsme.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
        #     ppto = wsme.cell(row=formula_ppto_prom_tri, column=c).coordinate
        #     wsme.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_revisado}:{c_fin_revisado})'
        #     revisado = wsme.cell(row=formula_revisado_prom_tri, column=c).coordinate
        #     wsme.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'
 
        row_count_memo += len(df_mo) + 3

        print("---------------------------------")
        print(f"Cuenta de fila de pestaña MEMO: {row_count_memo}")
        print("Escribiendo Total MEMO....")

        #Dataframe total MEMO
        row_ppto = total_memo_df.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2 + row_count_memo
        row_revisado = total_memo_df.count() + 1 + row_count_memo
        df_total_memo = (total_memo_df
                         .withColumn("TOTAL ANUAL",lit(""))
                         .withColumn("PROMEDIO ANUAL",lit(""))
                         .toPandas()
                        )
        df_total_memo.to_excel(writer, sheet_name="ESC-MEMO", startrow=row_count_memo,index=False, header=True)
        ws = writer.sheets["ESC-MEMO"]
        for r in range(row_count_memo + 2, row_count_memo + len(df_total_memo) + 2):
            ws.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            ws.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        # formula_ppto_prom_tri = row_count_mi + len(df) + 2
        # formula_revisado_prom_tri = row_count_mi + len(df) + 3
        # formula_alc_vs_ppto = row_count_mi + len(df) + 4

        # for c in (6,9,12,15):
        #     c_inicio_ppto = ws.cell(row = row_ppto, column= c - 2).coordinate
        #     c_fin_ppto = ws.cell(row = row_ppto, column= c).coordinate
        #     c_inicio_revisado = ws.cell(row = row_revisado, column= c - 3).coordinate
        #     c_fin_revisado = ws.cell(row = row_revisado, column= c).coordinate
        #     ws.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
        #     ppto = ws.cell(row=formula_ppto_prom_tri, column=c).coordinate
        #     ws.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
        #     revisado = ws.cell(row=formula_revisado_prom_tri, column=c).coordinate
        #     ws.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        #Dataframe total DACOMSA
        row_ppto = total_dacomsa_df.where(col("ANIO_FECHA_REGISTRO").contains("REAL")).count() + 2
        row_revisado = total_dacomsa_df.count() + 1
        df_total_dacomsa = (total_dacomsa_df
                           .withColumn("TOTAL ANUAL",lit(""))
                           .withColumn("PROMEDIO ANUAL",lit(""))
                           .toPandas()
                           )
        df_total_dacomsa.to_excel(writer, sheet_name="ESC-TOTAL", startrow=0,index=False, header=True)
        wst = writer.sheets["ESC-TOTAL"]
        for r in range(2, len(df_total_dacomsa) + 2):
            wst.cell(row=r, column=36).value = f'=SUM(D{r}:O{r})'
            wst.cell(row=r,column=37).value = f'=ROUND(AVERAGE(D{r}:O{r}),2)'
        formula_ppto_prom_tri = row_count_mi + len(df) + 2
        formula_revisado_prom_tri = row_count_mi + len(df) + 3
        formula_alc_vs_ppto = row_count_mi + len(df) + 4

        for c in (6,9,12,15):
            c_inicio_ppto = wst.cell(row = row_ppto, column= c - 2).coordinate
            c_fin_ppto = wst.cell(row = row_ppto, column= c).coordinate
            c_inicio_revisado = wst.cell(row = row_revisado, column= c - 2).coordinate
            c_fin_revisado = wst.cell(row = row_revisado, column= c).coordinate
            wst.cell(row=formula_ppto_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            ppto = wst.cell(row=formula_ppto_prom_tri, column=c).coordinate
            wst.cell(row=formula_revisado_prom_tri, column=c).value = f'=AVERAGE({c_inicio_ppto}:{c_fin_ppto})'
            revisado = wst.cell(row=formula_revisado_prom_tri, column=c).coordinate
            wst.cell(row=formula_alc_vs_ppto, column=c).value = f'=ROUND(({ppto}/{revisado})*100,2)'

        wst.cell(row=formula_ppto_prom_tri, column= 1).value = "PROMEDIO TRI PPTO"
        wst.cell(row=formula_revisado_prom_tri, column= 1).value = "PROMEDIO TRI REV"
        wst.cell(row=formula_alc_vs_ppto, column= 1).value = "ALCANCE"

        row_sumif_totales_memo_max = row_count_memo #- len(df_total_memo)
        
        max_mi = row_count_mi - len(df_total_dacomsa) - 1

        for c in range(4,36):
            celda_inicio = wst.cell(row = rango_sumif_max, column=c).coordinate
            celda_fin = wst.cell(row = max_mi , column=c).coordinate
            celda_inicio_memo = wst.cell(row = 2, column=c).coordinate
            celda_fin_memo = wst.cell(row = row_sumif_totales_memo_max, column=c).coordinate
            hoja_mi = "'ESC-MI'"
            hoja_me = "'ESC-MEMO'"
            wst.cell(row=row_revisado, column=c).value = f'=SUMIFS({hoja_mi}!{celda_inicio}:{celda_fin}, {hoja_mi}!$B${rango_sumif_max}:$B${max_mi}, "TOTAL", {hoja_mi}!$A${rango_sumif_max}:$A${max_mi}, "REVISADO") + SUMIFS({hoja_me}!{celda_inicio_memo}:{celda_fin_memo},{hoja_me}!$B$2:$B${row_sumif_totales_memo_max},"TOTAL",{hoja_me}!$A$2:$A${row_sumif_totales_memo_max},"PRONO-MKT")'

        print("TOTAL-DACOMSA")
        print(f"Celda ppto:{c_inicio_ppto}")
        print(f"Celda revisado:{c_inicio_revisado}")

    wb = load_workbook(temp_path)
    ws = wb["ESC-MI"]
    wsme = wb["ESC-MEMO"]
    wstot = wb["ESC-TOTAL"]
    
    for c in range(3, 40):
        for row in range(1, 1800):
            cell = ws.cell(row=row, column=c)
            cell.number_format = '0.00'  # Formato de número con dos decimales

    for c in range(3, 40):
        for row in range(1, 1800):
            cell = wsme.cell(row=row, column=c)
            cell.number_format = '0.00'  # Formato de número con dos decimales

    for c in range(3, 40):
        for row in range(1, 1800):
            cell = wstot.cell(row=row, column=c)
            cell.number_format = '0.00'  # Formato de número con dos decimales

    wb.save(temp_path)

    dbutils.fs.cp(f'file:{temp_path}', "/mnt/" + final_path)
    print(f"Archivo guardado en: {final_path}")